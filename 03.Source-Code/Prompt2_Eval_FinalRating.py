import os
import re
import json
import json5
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Evaluation of rating results obtained from intermediate information

# ======================== JSON File Processing Center ========================
class FileProcessor:
    """JSON File Processing Center"""

    @staticmethod
    def extract_json_from_content(content):
        """Extract JSON portion from mixed content"""
        markdown_match = re.search(r'```json\s*(.*?)\s*```', content, re.DOTALL)
        if markdown_match:
            return markdown_match.group(1).strip()

        json_obj_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_obj_match:
            return json_obj_match.group(0).strip()

        json_array_match = re.search(r'\[.*\]', content, re.DOTALL)
        if json_array_match:
            return json_array_match.group(0).strip()

        return content

    @staticmethod
    def fix_illegal_escapes(json_str):
        """Fix illegal escape sequences in JSON string"""
        illegal_escape_pattern = r'\\([^"\\/bfnrtu])'
        return re.sub(illegal_escape_pattern, r'\\\\\1', json_str)

    @staticmethod
    def structural_repair(json_str):
        """Intelligently repair JSON structure issues"""
        json_str = FileProcessor.fix_illegal_escapes(json_str)
        stack = []
        in_string = False
        result = []
        i = 0

        while i < len(json_str):
            char = json_str[i]
            if char == '\\' and i+1 < len(json_str):
                result.append(char + json_str[i+1])
                i += 2
                continue

            if char == '"':
                in_string = not in_string

            if not in_string:
                if char in ('{', '['):
                    stack.append(char)
                elif char == '}' and stack and stack[-1] == '{':
                    stack.pop()
                elif char == ']' and stack and stack[-1] == '[':
                    stack.pop()

            result.append(char)
            i += 1

        for left in reversed(stack):
            result.append('}' if left == '{' else ']')

        repaired = []
        in_string = False
        quote_count = 0
        for char in result:
            if char == '"' and (len(repaired) == 0 or repaired[-1] != '\\'):
                quote_count += 1
                in_string = not in_string
            repaired.append(char)

        if quote_count % 2 != 0:
            repaired.append('"')

        repaired_str = ''.join(repaired)
        last_char = repaired_str[-1] if repaired_str else ''
        if not re.search(r'[\]}\d"truefalsenull]$', last_char):
            if re.search(r':\s*[{\[]', repaired_str):
                repaired_str += ']}' if '{' in repaired_str else ']]'
            else:
                repaired_str += '}' if '{' in repaired_str else ']'

        return repaired_str

    @staticmethod
    def remove_json_comments(json_str):
        """Remove comments from JSON"""
        pattern = r'(\s*//.*?\n)|(/\*.*?\*/)'
        return re.sub(pattern, '', json_str, flags=re.DOTALL)

    @staticmethod
    def safe_json_parse(json_str):
        """JSON parsing with automatic repair"""
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            fixed = re.sub(r',\s*}(?=\s*})', '}', json_str)
            fixed = re.sub(r'[\x00-\x1F\x7F]', '', fixed)
            try:
                return json.loads(fixed)
            except:
                return json5.loads(fixed)

    @staticmethod
    def load_json(file_path):
        """Intelligently load JSON file"""
        encodings = ['utf-8-sig', 'utf-8', 'gb18030', 'latin-1']
        content = None

        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.read()
                break
            except:
                continue

        if not content:
            raise ValueError("All encoding attempts failed")

        json_content = FileProcessor.extract_json_from_content(content)
        cleaned_content = FileProcessor.remove_json_comments(json_content)
        repaired_content = FileProcessor.structural_repair(cleaned_content)

        try:
            return FileProcessor.safe_json_parse(repaired_content)
        except Exception as e:
            try:
                return FileProcessor.safe_json_parse(cleaned_content)
            except:
                raise ValueError(f"JSON parsing failed: {str(e)}")

    @staticmethod
    def save_extracted_json(json_data, output_path):
        """Save extracted JSON to file"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)

# ======================== Variant Analyzer ========================
class VariantAnalyzer:
    """Variant Analyzer"""

    @staticmethod
    def determine_strength_by_oddpath(odds_path, is_perfect_binary=None):
        """Determine evidence strength based on OddsPath value and conditions"""
        strength = "Supporting"
        if (odds_path < 0.0029) or (odds_path > 350):
            strength = "Very Strong"
        elif (0.0029 <= odds_path < 0.053) or (18.7 < odds_path <= 350):
            strength = "Strong"
        elif (0.053 <= odds_path < 0.23) or (4.3 < odds_path <= 18.7):
            strength = "Moderate"
        return strength

    @staticmethod
    def determine_evidence_strength(data):
        """Determine evidence strength for gene variant-disease association based on multi-level conditions"""
        # First level: Check if experimental method is approved
        if not VariantAnalyzer.evaluate_assay_validity_approved(data):
            return "No PS3/BS3"

        # Second level: Check if experiment includes valid controls and replicates
        if not VariantAnalyzer.evaluate_assay_validity_control(data):
            return "No PS3/BS3"

        # Third level: Check if experiment includes known pathogenic/benign variants
        if not VariantAnalyzer.evaluate_assay_contains_known_variants(data):
            return "Supporting"

        # Fourth level: Check if OddsPath can be calculated
        can_calculate_oddpath, odds_path, is_perfect_binary = VariantAnalyzer.calculate_oddpath(data)
        if not can_calculate_oddpath:
            # Cannot calculate OddsPath, count total pathogenic/benign variants
            pathogenic_count, benign_count = VariantAnalyzer.count_pathogenic_benign_variants(data)
            total_count = pathogenic_count + benign_count
            if total_count > 10:
                return "Moderate"
            else:
                return "Supporting"
        else:
            # Determine evidence strength based on OddsPath value and conditions
            return VariantAnalyzer.determine_strength_by_oddpath(odds_path, is_perfect_binary)

    @staticmethod
    def evaluate_assay_contains_known_variants(data):
        """Check if experiment includes known pathogenic/benign variants"""
        try:
            for assay in data["Experiment Method"]:
                # Process Validation controls P/LP field
                pathogenic_field = assay.get("Validation controls P/LP")
                if isinstance(pathogenic_field, dict):
                    has_pathogenic = pathogenic_field.get("Validation controls P/LP") == "Yes"
                elif pathogenic_field is None:
                    has_pathogenic = False
                else:
                    has_pathogenic = pathogenic_field == "Yes"

                # Process Validation controls B/LB field
                benign_field = assay.get("Validation controls B/LB")
                if isinstance(benign_field, dict):
                    has_benign = benign_field.get("Validation controls B/LB") == "Yes"
                elif benign_field is None:
                    has_benign = False
                else:
                    has_benign = benign_field == "Yes"

                if has_pathogenic or has_benign:
                    return True
            return False
        except (KeyError, TypeError):
            return False

    @staticmethod
    def calculate_oddpath(data):
        """Calculate oddpath value and determine condition (perfect binary or allows one indeterminate reading)"""
        pathogenic_count = 0
        benign_count = 0
        indeterminate_count = 0

        try:
            for assay in data["Experiment Method"]:
                # Get pathogenic variant count from Validation controls P/LP
                p_field = assay.get("Validation controls P/LP")
                p_counts = p_field.get("Counts", 0) if isinstance(p_field, dict) else 0
                pathogenic_count += int(p_counts) if str(p_counts).isdigit() else 0

                # Get benign variant count from Validation controls B/LB
                b_field = assay.get("Validation controls B/LB")
                b_counts = b_field.get("Counts", 0) if isinstance(b_field, dict) else 0
                benign_count += int(b_counts) if str(b_counts).isdigit() else 0

                # Count indeterminate conclusions
                readout = assay.get("Readout description")
                if isinstance(readout, list):
                    for result in readout:
                        if isinstance(result, dict) and result.get("Conclusion") == "Indeterminate":
                            indeterminate_count += 1

            # Calculate prior probability P1
            total_variants = pathogenic_count + benign_count
            if total_variants > 0:
                p1 = (pathogenic_count + 1) / (total_variants + 2)
                p2 = (pathogenic_count + 1) / (total_variants + 2)
                odds_path = (p2 * (1 - p1)) / ((1 - p2) * p1)
            else:
                return False, 0.0, True

            if indeterminate_count == 0:
                return True, odds_path, True
            elif indeterminate_count == 1:
                return True, odds_path, False
            else:
                return False, 0.0, True
        except ZeroDivisionError:
            return False, 0.0, True
        except (ValueError, TypeError):
            return False, 0.0, True

    @staticmethod
    def count_pathogenic_benign_variants(data):
        """Count the number of pathogenic and benign variants in the experiment"""
        pathogenic_count = 0
        benign_count = 0

        try:
            for assay in data["Experiment Method"]:
                readout = assay.get("Readout description")
                if not readout:
                    continue

                if isinstance(readout, list):
                    for result in readout:
                        if isinstance(result, dict):
                            conclusion = result.get("Conclusion")
                            if conclusion == "Abnormal":
                                pathogenic_count += 1
                            elif conclusion == "Normal":
                                benign_count += 1
                elif isinstance(readout, str):
                    # String format cannot distinguish different variants, skip counting
                    pass
        except (KeyError, TypeError):
            pass

        return pathogenic_count, benign_count

    @staticmethod
    def evaluate_assay_validity_approved(data):
        """Check if experimental method is Approved (at least one method marked as Yes)"""
        try:
            for assay in data.get("Experiment Method", []):
                # Process Approved assay field
                approved_field = assay.get("Approved assay")
                if approved_field is None:
                    continue

                if isinstance(approved_field, dict):
                    approved_value = approved_field.get("Approved assay", "No")
                else:
                    approved_value = approved_field

                if approved_value == "Yes":
                    return True
            return False
        except (KeyError, TypeError):
            return False

    @staticmethod
    def evaluate_assay_validity_control(data):
        """Evaluate if experiment includes valid controls and replicates"""
        try:
            for assay in data["Experiment Method"]:
                # Process Basic positive control
                basic_pos_control = assay.get("Basic positive control")
                if basic_pos_control is None:
                    has_basic_pos = False
                elif isinstance(basic_pos_control, dict):
                    has_basic_pos = basic_pos_control.get("Basic positive control") == "Yes"
                else:
                    has_basic_pos = basic_pos_control == "Yes"

                # Process Basic negative control
                basic_neg_control = assay.get("Basic negative control")
                if basic_neg_control is None:
                    has_basic_neg = False
                elif isinstance(basic_neg_control, dict):
                    has_basic_neg = basic_neg_control.get("Basic negative control") == "Yes"
                else:
                    has_basic_neg = basic_neg_control == "Yes"

                # Check biological replicates
                bio_replicates = assay.get("Biological replicates")
                if bio_replicates is None:
                    has_bio_reps = False
                elif isinstance(bio_replicates, dict):
                    has_bio_reps = bio_replicates.get("Biological replicates") == "Yes"
                else:
                    has_bio_reps = bio_replicates == "Yes"

                # Check technical replicates
                tech_replicates = assay.get("Technical replicates")
                if tech_replicates is None:
                    has_tech_reps = False
                elif isinstance(tech_replicates, dict):
                    has_tech_reps = tech_replicates.get("Technical replicates") == "Yes"
                else:
                    has_tech_reps = tech_replicates == "Yes"

                # Check basic controls (at least one positive or negative control)
                has_basic_control = has_basic_pos or has_basic_neg
                # Check replicates (at least one biological or technical replicate)
                has_replicates = has_bio_reps or has_tech_reps

                if has_basic_control and has_replicates:
                    return True
            return False
        except (KeyError, TypeError):
            return False

    @staticmethod
    def read_json_file(file_path):
        """Read JSON file and return data"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return json.load(file)
        except (FileNotFoundError, json.JSONDecodeError, Exception) as e:
            print(f"Error: {e}")
            return None

    @staticmethod
    def analyze_variants_evidence(data):
        """Analyze pathogenicity evidence strength for each variant"""
        variants = {}
        # Extract all variant information
        for gene_info in data.get("Variants Include", []):
            for var in gene_info.get("variants", []):
                hgvs = var.get("HGVS")
                if hgvs:
                    variants[hgvs] = {
                        "HGVS": hgvs,
                        "Protein Change": var.get("Protein Change"),
                        "Descriptions": [],
                        "Conclusions": [],
                        "is_pathogenic": None,
                        "evidence_strength": "No PS3/BS3"
                    }

        # Collect conclusions for each variant across different experiments
        for assay in data.get("Experiment Method", []):
            assay_name = assay.get("Assay Method", "Unknown Assay")
            readout = assay.get("Readout description")

            # Case 1: Readout description is a list
            if isinstance(readout, list):
                for result in readout:
                    if not isinstance(result, dict):
                        continue

                    variant_hgvs = result.get("Variant")
                    conclusion = result.get("Conclusion")
                    if variant_hgvs in variants and conclusion:
                        variants[variant_hgvs]["Conclusions"].append({
                            "assay": assay_name,
                            "conclusion": conclusion,
                            "molecular_effect": result.get("Molecular Effect"),
                            "description": result.get("Result Description")
                        })

            # Case 2: Has Readout details
            elif "Readout details" in assay:
                readout_details = assay.get("Readout details", {})
                for gene, gene_data in readout_details.items():
                    for var_desc, description in gene_data.items():
                        # Find matching variant
                        for hgvs, variant in variants.items():
                            if var_desc == hgvs or (variant["Protein Change"] and
                                                    var_desc == f"{variant['Protein Change'].get('ref', '')}"
                                                                f"{variant['Protein Change'].get('position', '')}"
                                                                f"{variant['Protein Change'].get('alt', '')}"):
                                variants[hgvs]["Conclusions"].append({
                                    "assay": assay_name,
                                    "conclusion": "Abnormal" if "Increased" in description or "Reduced" in description else "Unknown",
                                    "molecular_effect": description,
                                    "description": description
                                })

            # Case 3: Readout description is a string
            elif isinstance(readout, str):
                # Attempt to extract conclusions from description
                for hgvs in variants:
                    if hgvs in readout:
                        variants[hgvs]["Conclusions"].append({
                            "assay": assay_name,
                            "conclusion": "Abnormal" if "Increased" in readout or "Reduced" in readout else "Normal",
                            "molecular_effect": readout,
                            "description": readout
                        })

        # Evaluate pathogenicity and evidence strength for each variant
        for hgvs, var_data in variants.items():
            conclusions = [concl["conclusion"] for concl in var_data["Conclusions"]]

            # Determine pathogenicity
            if all(concl == "N.D." for concl in conclusions):
                var_data["is_pathogenic"] = None
            else:
                var_data["is_pathogenic"] = any(concl == "Abnormal" for concl in conclusions)

            # Create subset data for this variant - use original exact match method
            variant_data = {
                "Article Info": data.get("Article Info", {}),
                "Experiment Method": []
            }

            # Select experiments using original exact match method
            for assay in data.get("Experiment Method", []):
                readout = assay.get("Readout description")
                if not readout:
                    continue

                found = False
                # Process list format Readout
                if isinstance(readout, list):
                    for result in readout:
                        if isinstance(result, dict) and result.get("Variant") == hgvs:
                            found = True
                            break
                # Process Readout details format
                elif "Readout details" in assay:
                    readout_details = assay.get("Readout details", {})
                    for gene, gene_data in readout_details.items():
                        if hgvs in gene_data:
                            found = True
                            break
                # Process string format
                elif isinstance(readout, str) and hgvs in readout:
                    found = True

                if found:
                    variant_data["Experiment Method"].append(assay)

            # Calculate evidence strength using existing logic
            if variant_data["Experiment Method"]:
                strength = VariantAnalyzer.determine_evidence_strength(variant_data)
                var_data["evidence_strength"] = strength
            else:
                var_data["evidence_strength"] = "No PS3/BS3"

        return list(variants.values())

    @staticmethod
    def save_variant_analysis_to_excel(data, filename="variant_analysis.xlsx"):
        """Save variant analysis results to Excel file"""
        variant_analyses = VariantAnalyzer.analyze_variants_evidence(data)

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Variant Analysis Results"

        # Write headers
        headers = ["Variant", "Pathogenicity", "Evidence Strength", "Experiment Conclusions", "Molecular Effect", "Description"]
        ws.append(headers)

        # Write data
        for analysis in variant_analyses:
            protein_change = analysis["Protein Change"]
            change_desc = f"{protein_change.get('ref', '?')}{protein_change.get('position', '?')}{protein_change.get('alt', '?')}" if protein_change else "?"

            pathogenic_text = "Unknown" if analysis["is_pathogenic"] is None else (
                "Pathogenic" if analysis["is_pathogenic"] else "Benign")

            # Combine multiple experiment conclusions
            assay_conclusions = []
            molecular_effects = []
            descriptions = []

            for concl in analysis["Conclusions"]:
                assay_conclusions.append(f"{concl['assay']}: {concl['conclusion']}")
                molecular_effects.append(concl["molecular_effect"] or "")
                descriptions.append(concl["description"] or "")

            # Write main row
            main_row = [
                f"{analysis['HGVS']} ({change_desc})",
                pathogenic_text,
                analysis["evidence_strength"],
                "\n".join(assay_conclusions),
                "\n".join(molecular_effects),
                "\n".join(descriptions)
            ]
            ws.append(main_row)

        # Set column widths and wrap text
        for col in range(1, 7):
            col_letter = get_column_letter(col)
            if col in [4, 5, 6]:  # Experiment Conclusions, Molecular Effect, Description columns
                ws.column_dimensions[col_letter].width = 50
            else:
                ws.column_dimensions[col_letter].width = 20

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Set header row style
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save file
        wb.save(filename)
        print(f"Variant analysis results saved to {filename}")
        return filename

    @staticmethod
    def print_variant_analysis(data):
        """Print variant analysis results"""
        variant_analyses = VariantAnalyzer.analyze_variants_evidence(data)

        print(f"\nPMID: {data['Article Info'].get('PMID', 'Unknown')}")
        print("Variant Evidence Strength Analysis Results:")
        for analysis in variant_analyses:
            protein_change = analysis["Protein Change"]
            change_desc = f"{protein_change.get('ref', '?')}{protein_change.get('position', '?')}{protein_change.get('alt', '?')}" if protein_change else "?"

            pathogenic_text = "Unknown" if analysis["is_pathogenic"] is None else (
                "Pathogenic" if analysis["is_pathogenic"] else "Benign")

            print(f"\nVariant: {analysis['HGVS']} ({change_desc})")
            print(f"Pathogenicity: {pathogenic_text}")
            print(f"Evidence Strength: {analysis['evidence_strength']}")

            print("Experiment Conclusions:")
            for i, concl in enumerate(analysis["Conclusions"], 1):
                print(f"  {i}. {concl['assay']}: {concl['conclusion']}")
                print(f"     - Molecular Effect: {concl['molecular_effect']}")
                print(f"     - Description: {concl['description']}")

# ======================== Main Processing Flow ========================
def extract_pmid(filename):
    """Extract PMID from filename"""
    match = re.search(r'(\d{8})', filename)
    return match.group(1) if match else None

def find_corresponding_files(standard_dir, model_dir):
    """Find corresponding PMID files in two directories (optimized version)"""
    standard_files = {}
    model_files = {}

    # Process standard directory - unchanged
    for root, _, files in os.walk(standard_dir):
        for file in files:
            pmid = extract_pmid(file)
            if pmid and file.endswith('.json'):
                standard_files[pmid] = os.path.join(root, file)

    # Process model output directory - add support for _result format
    for root, _, files in os.walk(model_dir):
        for file in files:
            pmid = extract_pmid(file)
            if pmid:
                # Handle multiple formats:
                # 1. JSON files named directly with PMID (e.g., 10888878.json)
                # 2. Files ending with _result (e.g., 10888878_result.txt)
                # 3. Files ending with 01 (e.g., 10888878_gemma3_01.txt)
                # 4. Other files containing PMID
                if (file.endswith('.json') and file.startswith(pmid)) or \
                        (file.endswith('_result.txt') or file.endswith('_result.json')) or \
                        (file.endswith('01.txt') or file.endswith('01.json')) or \
                        (pmid in file and (file.endswith('.txt') or file.endswith('.json'))):
                    model_files[pmid] = os.path.join(root, file)

    return standard_files, model_files

def extract_variant_change(description):
    """Extract amino acid change portion from variant description"""
    match = re.search(r'\(([A-Z]\d+[A-Z])\)', description)
    return match.group(1) if match else None

def extract_amino_acid_change(description):
    """Extract amino acid change portion from variant description (content in parentheses)"""
    # Enhanced regex to match stop codon formats (Stop/Ter/*)
    pattern = r'\(([A-Za-z]{1,3}\d*[A-Za-z*]*\d*(?:[A-Za-z]*|\d*[A-Za-z*]*)(?:[A-Za-z*]+\d*|\d+[A-Za-z*]+)*)\)'
    match = re.search(pattern, description)

    # If no match, try a more general pattern
    if not match:
        # Try matching combinations of numbers and letters
        pattern = r'\(([\w\d_/*]+)\)'
        match = re.search(pattern, description)

    return match.group(1).strip().upper() if match else None

def compare_variants(model_results, standard_results):
    """Compare variant analysis between model output and standard results (detailed multi-dimensional comparison)"""
    # Variant comparison
    std_variants = {}
    model_variants = {}

    for result in standard_results:
        aa_change = extract_amino_acid_change(result['Variant'])
        if aa_change:
            # Normalize stop codon format
            normalized_aa = re.sub(r'(STOP|TER)$', '*', aa_change)
            std_variants[normalized_aa] = {
                'pathogenic': result['Pathogenicity'],
                'strength': result['Evidence Strength']
            }

    for result in model_results:
        aa_change = extract_amino_acid_change(result['Variant'])
        if aa_change:
            # Normalize stop codon format
            normalized_aa = re.sub(r'(STOP|TER)$', '*', aa_change)
            model_variants[normalized_aa] = {
                'pathogenic': result['Pathogenicity'],
                'strength': result['Evidence Strength']
            }

    # 1. Variant dimension: Compare if amino acid changes are identical
    variant_matches = [aa for aa in model_variants if aa in std_variants]
    variant_correct = len(variant_matches)

    # 2. Pathogenicity dimension: Compare if pathogenicity matches for identical variants
    pathogenicity_matches = []
    for aa in variant_matches:
        if model_variants[aa]['pathogenic'] == std_variants[aa]['pathogenic']:
            pathogenicity_matches.append(aa)

    pathogenicity_correct = len(pathogenicity_matches)

    # 3. Evidence strength dimension: Compare if evidence strength matches for identical variants and pathogenicity
    strength_matches = []
    for aa in pathogenicity_matches:
        if model_variants[aa]['strength'] == std_variants[aa]['strength']:
            strength_matches.append(aa)

    strength_correct = len(strength_matches)

    return {
        'std_variants_count': len(std_variants),
        'model_variants_count': len(model_variants),
        'variant_correct': variant_correct,
        'pathogenicity_correct': pathogenicity_correct,
        'strength_correct': strength_correct
    }

def process_pmid_results(pmid, extracted_dir, model_results_dir):
    """Process results comparison for a single PMID (return multi-dimensional statistics)"""
    standard_json_path = os.path.join(extracted_dir, 'standard', f"{pmid}.json")
    model_json_path = os.path.join(extracted_dir, 'model', f"{pmid}.json")

    # Process standard results
    standard_data = FileProcessor.load_json(standard_json_path)
    standard_analysis = VariantAnalyzer.analyze_variants_evidence(standard_data)

    # Process model results
    model_data = FileProcessor.load_json(model_json_path)
    model_analysis = VariantAnalyzer.analyze_variants_evidence(model_data)

    # Convert results to format required for comparison
    standard_results = []
    for analysis in standard_analysis:
        protein_change = analysis["Protein Change"]
        change_desc = f"{protein_change.get('ref', '?')}{protein_change.get('position', '?')}{protein_change.get('alt', '?')}" if protein_change else "?"
        pathogenic_text = "Unknown" if analysis["is_pathogenic"] is None else ("Pathogenic" if analysis["is_pathogenic"] else "Benign")
        standard_results.append({
            "Variant": f"{analysis['HGVS']} ({change_desc})",
            "Pathogenicity": pathogenic_text,
            "Evidence Strength": analysis["evidence_strength"]
        })

    model_results = []
    for analysis in model_analysis:
        protein_change = analysis["Protein Change"]
        change_desc = f"{protein_change.get('ref', '?')}{protein_change.get('position', '?')}{protein_change.get('alt', '?')}" if protein_change else "?"
        pathogenic_text = "Unknown" if analysis["is_pathogenic"] is None else ("Pathogenic" if analysis["is_pathogenic"] else "Benign")
        model_results.append({
            "Variant": f"{analysis['HGVS']} ({change_desc})",
            "Pathogenicity": pathogenic_text,
            "Evidence Strength": analysis["evidence_strength"]
        })

    # Compare results and get multi-dimensional statistics
    comparison_result = compare_variants(model_results, standard_results)

    # Save model analysis results to Excel
    model_result_excel = os.path.join(model_results_dir, f"{pmid}.xlsx")
    os.makedirs(os.path.dirname(model_result_excel), exist_ok=True)
    VariantAnalyzer.save_variant_analysis_to_excel(model_data, model_result_excel)

    return {
        'PMID': pmid,
        'Variant Dimension': {
            'Standard Count': comparison_result['std_variants_count'],
            'Model Count': comparison_result['model_variants_count'],
            'Correct Count': comparison_result['variant_correct']
        },
        'Pathogenicity Dimension': {
            'Standard Count': comparison_result['variant_correct'],  # Based on correct variant count
            'Model Count': comparison_result['variant_correct'],  # Based on correct variant count
            'Correct Count': comparison_result['pathogenicity_correct']
        },
        'Evidence Strength Dimension': {
            'Standard Count': comparison_result['pathogenicity_correct'],  # Based on correct pathogenicity
            'Model Count': comparison_result['pathogenicity_correct'],  # Based on correct pathogenicity
            'Correct Count': comparison_result['strength_correct']
        }
    }

def main():
    """Main processing flow (updated to support multi-dimensional statistics)"""
    start_time = datetime.now()
    print(f"Processing started: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

    # Configuration paths (unchanged)
    # Modify the path according to the actual situation
    standard_dir = r'../ref_json'
    model_dir = r'../result/01llama_70b_textRAG'
    extracted_dir = r'../local_01llama_70b_textRAG/extracted_json'
    model_results_dir = r'../local_01llama_70b_textRAG/model_results_excel'
    final_report = r'../local_01llama_70b_textRAG/final_results.txt'

    # Ensure output directories exist
    os.makedirs(extracted_dir, exist_ok=True)
    os.makedirs(os.path.join(extracted_dir, 'standard'), exist_ok=True)
    os.makedirs(os.path.join(extracted_dir, 'model'), exist_ok=True)
    os.makedirs(model_results_dir, exist_ok=True)

    # Find corresponding PMID files
    standard_files, model_files = find_corresponding_files(standard_dir, model_dir)
    print(f"Found {len(standard_files)} standard files, {len(model_files)} model files")

    # Extract JSON files
    all_results = []
    processed_pmids = set()

    # Process standard files
    for pmid, standard_path in standard_files.items():
        try:
            json_data = FileProcessor.load_json(standard_path)
            output_path = os.path.join(extracted_dir, 'standard', f"{pmid}.json")
            FileProcessor.save_extracted_json(json_data, output_path)
            processed_pmids.add(pmid)
            print(f"Standard file processed successfully: {pmid}")
        except Exception as e:
            print(f"Standard file processing failed - PMID: {pmid}, Error: {str(e)}")

    # Process model files
    for pmid, model_path in model_files.items():
        try:
            json_data = FileProcessor.load_json(model_path)
            output_path = os.path.join(extracted_dir, 'model', f"{pmid}.json")
            FileProcessor.save_extracted_json(json_data, output_path)
            processed_pmids.add(pmid)
            print(f"Model file processed successfully: {pmid}")
        except Exception as e:
            print(f"Model file processing failed - PMID: {pmid}, Error: {str(e)}")

    # Initialize totals
    total_variant = {'std_count': 0, 'model_count': 0, 'correct_count': 0}
    total_pathogenicity = {'std_count': 0, 'model_count': 0, 'correct_count': 0}
    total_strength = {'std_count': 0, 'model_count': 0, 'correct_count': 0}
    article_count = 0

    # Process results comparison for each PMID
    for pmid in processed_pmids:
        try:
            result = process_pmid_results(pmid, extracted_dir, model_results_dir)
            all_results.append(result)
            article_count += 1

            # Accumulate variant dimension totals
            var_data = result['Variant Dimension']
            total_variant['std_count'] += var_data['Standard Count']
            total_variant['model_count'] += var_data['Model Count']
            total_variant['correct_count'] += var_data['Correct Count']

            # Accumulate pathogenicity dimension totals
            path_data = result['Pathogenicity Dimension']
            total_pathogenicity['std_count'] += path_data['Standard Count']
            total_pathogenicity['model_count'] += path_data['Model Count']
            total_pathogenicity['correct_count'] += path_data['Correct Count']

            # Accumulate evidence strength dimension totals
            strength_data = result['Evidence Strength Dimension']
            total_strength['std_count'] += strength_data['Standard Count']
            total_strength['model_count'] += strength_data['Model Count']
            total_strength['correct_count'] += strength_data['Correct Count']

            print(f"PMID {pmid} processing completed: ")
            print(f"  Variant: Standard Count={var_data['Standard Count']}, Model Count={var_data['Model Count']}, Correct Count={var_data['Correct Count']}")
            print(f"  Pathogenicity: Standard Count={path_data['Standard Count']}, Model Count={path_data['Model Count']}, Correct Count={path_data['Correct Count']}")
            print(f"  Evidence Strength: Standard Count={strength_data['Standard Count']}, Model Count={strength_data['Model Count']}, Correct Count={strength_data['Correct Count']}")
        except Exception as e:
            print(f"PMID {pmid} result comparison failed: {str(e)}")

    # Save final results report (TXT format)
    with open(final_report, 'w', encoding='utf-8') as f:
        # Write header
        f.write("PMID\tVariant_Standard_Count\tVariant_Model_Count\tVariant_Correct_Count\tPathogenicity_Standard_Count\tPathogenicity_Model_Count\tPathogenicity_Correct_Count\tEvidence_Strength_Standard_Count\tEvidence_Strength_Model_Count\tEvidence_Strength_Correct_Count\n")

        # Write results for each article
        for result in all_results:
            var = result['Variant Dimension']
            path = result['Pathogenicity Dimension']
            strength = result['Evidence Strength Dimension']

            f.write(f"{result['PMID']}\t{var['Standard Count']}\t{var['Model Count']}\t{var['Correct Count']}\t")
            f.write(f"{path['Standard Count']}\t{path['Model Count']}\t{path['Correct Count']}\t")
            f.write(f"{strength['Standard Count']}\t{strength['Model Count']}\t{strength['Correct Count']}\n")

        # Write totals
        f.write("\nTotals:\n")
        f.write(f"Total Articles: {article_count}\n")
        f.write(f"Variant Totals - Standard Count: {total_variant['std_count']}, Model Count: {total_variant['model_count']}, Correct Count: {total_variant['correct_count']}\n")
        f.write(f"Pathogenicity Totals - Standard Count: {total_pathogenicity['std_count']}, Model Count: {total_pathogenicity['model_count']}, Correct Count: {total_pathogenicity['correct_count']}\n")
        f.write(f"Evidence Strength Totals - Standard Count: {total_strength['std_count']}, Model Count: {total_strength['model_count']}, Correct Count: {total_strength['correct_count']}\n")

    print(f"Final results saved to: {final_report}")

    end_time = datetime.now()
    duration = end_time - start_time
    print(f"Processing completed: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total duration: {duration.total_seconds():.2f} seconds")

if __name__ == "__main__":
    main()