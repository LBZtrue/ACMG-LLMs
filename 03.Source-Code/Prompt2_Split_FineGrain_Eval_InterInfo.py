import fnmatch
import json
import re
import os
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension

# Fine grained evaluation of intermediate information extracted from LLMs

class Config:
    """Global Configuration Center"""

    # Comparison Configuration
    EXACT_MATCH_FIELDS = {
        "Variants Include.Gene",
        "Variants Include.variants.cDNA Change.transcript",
        "Variants Include.variants.cDNA Change.ref",
        "Variants Include.variants.cDNA Change.alt",
        "Variants Include.variants.cDNA Change.position",
        "Variants Include.variants.Protein Change.ref",
        "Variants Include.variants.Protein Change.alt",
        "Variants Include.variants.Protein Change.position",
        "Variants Include.variants.Description in input context"
    }

    SKIP_FIELDS = {
        "Variants Include.variants.Description in input context"
    }

    BOOLEAN_FIELDS = {}

    FIELD_GROUPS = {
        "Variants Include.variants.cDNA Change.ref": "cDNA Change",
        "Variants Include.variants.cDNA Change.alt": "cDNA Change",
        "Variants Include.variants.cDNA Change.position": "cDNA Change",
        "Variants Include.variants.Protein Change.ref": "Protein Change",
        "Variants Include.variants.Protein Change.alt": "Protein Change",
        "Variants Include.variants.Protein Change.position": "Protein Change",
    }

    AMINO_ACID_MAP = {
        'ala': 'A', 'alanine': 'A', 'A': 'A',
        'arg': 'R', 'arginine': 'R', 'R': 'R',
        'asn': 'N', 'asparagine': 'N', 'N': 'N',
        'asp': 'D', 'aspartic acid': 'D', 'D': 'D',
        'cys': 'C', 'cysteine': 'C', 'C': 'C',
        'gln': 'Q', 'glutamine': 'Q', 'Q': 'Q',
        'glu': 'E', 'glutamic acid': 'E', 'E': 'E',
        'gly': 'G', 'glycine': 'G', 'G': 'G',
        'his': 'H', 'histidine': 'H', 'H': 'H',
        'ile': 'I', 'isoleucine': 'I', 'I': 'I',
        'leu': 'L', 'leucine': 'L', 'L': 'L',
        'lys': 'K', 'lysine': 'K', 'K': 'K',
        'met': 'M', 'methionine': 'M', 'M': 'M',
        'phe': 'F', 'phenylalanine': 'F', 'F': 'F',
        'pro': 'P', 'proline': 'P', 'P': 'P',
        'ser': 'S', 'serine': 'S', 'S': 'S',
        'thr': 'T', 'threonine': 'T', 'T': 'T',
        'trp': 'W', 'tryptophan': 'W', 'W': 'W',
        'tyr': 'Y', 'tyrosine': 'Y', 'Y': 'Y',
        'val': 'V', 'valine': 'V', 'V': 'V'
    }

    NUCLEIC_ACID_MAP = {
        'adenine': 'A', 'A': 'A',
        'thymine': 'T', 'T': 'T',
        'cytosine': 'C', 'C': 'C',
        'guanine': 'G', 'G': 'G',
        'uracil': 'U', 'U': 'U'
    }

    # Excel Template Configuration
    LEFT_COLUMNS = 5  # Number of fixed left columns
    HEADER_ROW = 2
    FIELD_START_ROW = 5
    PMID_COL = 6  # Starting column for publication metrics

    FIELD_STRUCTURE = [
        ["Variants Include", "Gene", "", "Number matching reference results", ""],
        ["", "variants", "cDNA Change", "transcript", "Number matching reference results"],
        ["", "", "", "ref,alt,position", "Number matching reference results"],
        ["", "", "Protein Change", "ref,alt,position", "Number matching reference results"],
        ["", "", "Description in input context", "", "Length/Number matching original text"],
    ]

    INDICATOR_HEADERS = [
        '1 Standard Total', '1 Total Output', '1 Correct Count', '4 False Assertions', '5 Logical Conflicts',
        '2 Field Omissions', '6 Consistency Rate', '7 Processing Time', '9 Final Rating Consistency'
    ]

    FIELD_MAPPING = {
        "Variants Include.Gene": 4,
        "Variants Include.variants.cDNA Change.transcript": 6,
        "Variants Include.variants.cDNA Change.ref": 7,
        "Variants Include.variants.cDNA Change.alt": 7,
        "Variants Include.variants.cDNA Change.position": 7,
        "Variants Include.variants.Protein Change.ref": 8,
        "Variants Include.variants.Protein Change.alt": 8,
        "Variants Include.variants.Protein Change.position": 8,
        "Variants Include.variants.Description in input context": 9,
    }

    OMISSION_FIELDS = {
        "Variants Include.variants.Protein Change.ref",
        "Variants Include.variants.Protein Change.alt",
        "Variants Include.variants.Protein Change.position"
    }


class Step1_DataComparator:
    def __init__(self, std_json):
        self.std_json = std_json
        self.field_metrics = defaultdict(lambda: {
            'std_count': 0,        # 1 Standard Total
            'model_count': 0,      # 1 Total Output
            'correct': 0,         # 1 Correct Count
            'false_assert': 0,    # 4 False Assertions
            'logical_conflict': 0, # 5 Logical Conflicts
            'field_omissions': 0,  # 2 Field Omissions
            'correct_details': [], #
            'std_values': [],     # Added: Store all standard field values
            'model_values': [],   # Added: Store all model output values
            'matched_std_values': set()  # Added: Store matched standard values
        })

    def compare(self, model_json):
        self._traverse_node(self.std_json, model_json, is_std=True)
        self._traverse_node(model_json, self.std_json, is_std=False)
        self._process_field_groups()
        return self._calculate_metrics()

    def _traverse_node(self, current_node, compare_node, is_std, path=[]):
        current_path = '.'.join(path)
        print(f"[DEBUG] Traversing path: {current_path}, is_std={is_std}, type: {type(current_node)}")

        if isinstance(current_node, dict):
            for key in current_node:
                new_path = path + [key]
                compare_value = compare_node.get(key, None) if isinstance(compare_node, dict) else None
                self._traverse_node(current_node[key], compare_value, is_std, new_path)
        elif isinstance(current_node, list):
            for idx, item in enumerate(current_node):
                compare_item = compare_node[idx] if (isinstance(compare_node, list) and idx < len(compare_node)) else None
                self._traverse_node(item, compare_item, is_std, path.copy())
        else:
            if is_std:
                self._process_leaf_node(current_node, compare_node, is_std, path)
            else:
                # Special handling for flattened model output structure, mapping to standard field paths
                if path == ['Gene']:
                    mapped_path = ['Variants Include', 'Gene']
                elif path == ['variants', 'cDNA Change', 'transcript']:
                    mapped_path = ['Variants Include', 'variants', 'cDNA Change', 'transcript']
                elif path == ['variants', 'cDNA Change', 'ref']:
                    mapped_path = ['Variants Include', 'variants', 'cDNA Change', 'ref']
                elif path == ['variants', 'cDNA Change', 'alt']:
                    mapped_path = ['Variants Include', 'variants', 'cDNA Change', 'alt']
                elif path == ['variants', 'cDNA Change', 'position']:
                    mapped_path = ['Variants Include', 'variants', 'cDNA Change', 'position']
                elif path == ['variants', 'Protein Change', 'ref']:
                    mapped_path = ['Variants Include', 'variants', 'Protein Change', 'ref']
                elif path == ['variants', 'Protein Change', 'alt']:
                    mapped_path = ['Variants Include', 'variants', 'Protein Change', 'alt']
                elif path == ['variants', 'Protein Change', 'position']:
                    mapped_path = ['Variants Include', 'variants', 'Protein Change', 'position']
                elif path == ['variants', 'Description in input context']:
                    mapped_path = ['Variants Include', 'variants', 'Description in input context']
                else:
                    mapped_path = path.copy()  # Ensure a new list object is created

                # Ensure compare_node is the model value
                model_val = current_node  # Get the value of the current model node
                self._process_leaf_node(compare_node, model_val, is_std, mapped_path)

    def _process_leaf_node(self, std_val, model_val, is_std, path):
        field_path = '.'.join(path)
        metrics = self.field_metrics[field_path]

        print(f"[DEBUG] Processing field: {field_path}, is_std={is_std}")
        print(f"[DEBUG] Standard value: {std_val}, Model value: {model_val}")

        if field_path in Config.SKIP_FIELDS:
            if is_std:
                metrics['std_values'].append(str(std_val))
                metrics['std_count'] += 1
            else:
                model_str = str(model_val) if model_val is not None else ""
                metrics['model_values'].append(model_str)
                metrics['model_count'] += 1
                std_content = str(std_val).lower()
                model_content = model_str.lower()
                if std_content == model_content:
                    metrics['correct'] += 1
                    metrics['correct_details'].append(f"Standard: {std_val}, Output: {model_val}")
            return

        if field_path in Config.EXACT_MATCH_FIELDS:
            if is_std:
                clean_val = re.sub(r'[^a-zA-Z0-9]', '', str(std_val)).lower()
                metrics['std_values'].append(clean_val)
                metrics['std_count'] += 1
            else:
                model_clean = re.sub(r'[^a-zA-Z0-9]', '', str(model_val)).lower() if model_val else ''
                metrics['model_values'].append(model_clean)
                metrics['model_count'] += 1
                if metrics['std_values'] and model_clean in metrics['std_values']:
                    metrics['correct'] += 1
                else:
                    metrics['false_assert'] += 1
            return

        if field_path in Config.FIELD_GROUPS:
            if is_std:
                metrics['std_values'].append(str(std_val))
                metrics['std_count'] += 1
            else:
                metrics['model_values'].append(str(model_val))
                metrics['model_count'] += 1
            return

    def _process_field_groups(self):
        group_metrics = defaultdict(lambda: {
            'std_variants': set(),
            'model_variants': set(),
            'std_count': 0,
            'model_count': 0,
            'correct': 0,
            'false_assert': 0
        })

        for group_name in {"cDNA Change", "Protein Change"}:
            group_fields = [f for f in Config.FIELD_GROUPS if Config.FIELD_GROUPS[f] == group_name]
            ref_field, alt_field, pos_field = group_fields[0], group_fields[1], group_fields[2]

            std_refs = self.field_metrics[ref_field]['std_values']
            std_alts = self.field_metrics[alt_field]['std_values']
            std_positions = self.field_metrics[pos_field]['std_values']

            max_len = max(len(std_refs), len(std_alts), len(std_positions))
            valid_std = []

            for i in range(max_len):
                ref = std_refs[i] if i < len(std_refs) else ""
                alt = std_alts[i] if i < len(std_alts) else ""
                pos = std_positions[i] if i < len(std_positions) else ""

                valid_std.append((ref, alt, pos))

            group_metrics[group_name]['std_variants'] = set(valid_std)
            group_metrics[group_name]['std_count'] = len(valid_std)

            model_refs = self.field_metrics[ref_field]['model_values']
            model_alts = self.field_metrics[alt_field]['model_values']
            model_positions = self.field_metrics[pos_field]['model_values']

            max_len = max(len(model_refs), len(model_alts), len(model_positions))
            valid_model = []

            for i in range(max_len):
                ref = model_refs[i] if i < len(model_refs) else ""
                alt = model_alts[i] if i < len(model_alts) else ""
                pos = model_positions[i] if i < len(model_positions) else ""

                valid_model.append((ref, alt, pos))

            group_metrics[group_name]['model_variants'] = set(valid_model)
            group_metrics[group_name]['model_count'] = len(valid_model)

            correct = sum(1 for m in valid_model if m in valid_std)
            false_assert = len([m for m in valid_model if m not in valid_std])

            group_metrics[group_name]['correct'] = correct
            group_metrics[group_name]['false_assert'] = false_assert

        for field_path in Config.FIELD_GROUPS:
            group_name = Config.FIELD_GROUPS[field_path]
            self.field_metrics[field_path].update({
                'std_count': group_metrics[group_name]['std_count'],
                'model_count': group_metrics[group_name]['model_count'],
                'correct': group_metrics[group_name]['correct'],
                'false_assert': group_metrics[group_name]['false_assert']
            })

    def _calculate_metrics(self):
        total = defaultdict(int)
        processed_groups = set()

        for field_path, metrics in self.field_metrics.items():
            if field_path in Config.FIELD_GROUPS:
                group_name = Config.FIELD_GROUPS[field_path]
                if group_name in processed_groups:
                    continue

                group_fields = [f for f in Config.FIELD_GROUPS if Config.FIELD_GROUPS[f] == group_name]
                main_field = group_fields[0]
                main_metrics = self.field_metrics[main_field]

                total['std_total'] += main_metrics['std_count']
                total['model_total'] += main_metrics['model_count']
                total['correct_total'] += main_metrics['correct']
                total['false_assert_total'] += main_metrics['false_assert']

                processed_groups.add(group_name)
                continue

            total['std_total'] += metrics['std_count']
            total['model_total'] += metrics['model_count']
            total['correct_total'] += metrics['correct']
            total['false_assert_total'] += metrics['false_assert']

        return {
            'std_total': total['std_total'],
            'model_total': total['model_total'],
            'correct_total': total['correct_total'],
            'false_assert_total': total['false_assert_total'],
            'field_metrics': self.field_metrics
        }


class Step2_TemplateBuilder:
    @staticmethod
    def create_template(output_path):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        if os.path.exists(output_path):
            return load_workbook(output_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Results"

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 8
        for col in range(6, 50):
            ws.column_dimensions[get_column_letter(col)].width = 12

        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:D1')
        ws.merge_cells('E1:E3')

        ws['A1'] = "Model Name:"
        ws['C1'] = "Deployment Method"
        ws['E1'] = "Metrics"

        ws.merge_cells('A2:D3')
        ws['A2'] = "1 Information Extraction Accuracy 2 Information Completeness 3 Classification Performance 4 Generation Quality (Hallucination) 5 Interpretability 6 Consistency 7 Efficiency 8 Scalability 9 Clinical Utility"

        field_start_row = 4
        for idx, row in enumerate(Config.FIELD_STRUCTURE, start=field_start_row):
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws.append([''] * 5 + ['Per Column', 'Summary'])
        ws.freeze_panes = "F4"

        wb.save(output_path)
        return wb


class Step3_ResultFiller:
    @staticmethod
    def fill_results(wb, output_path, pmid, analysis_result, qa_time):
        ws = wb.active

        start_col = ws.max_column + 1 if ws.max_column > Config.LEFT_COLUMNS else Config.PMID_COL
        start_row = Config.HEADER_ROW

        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+len(Config.INDICATOR_HEADERS)-1)
        pmid_cell = ws.cell(row=start_row, column=start_col, value=f"PMID_{pmid}")
        pmid_cell.alignment = Alignment(horizontal='center', vertical='center')

        title_row = start_row + 1
        for col_idx, header in enumerate(Config.INDICATOR_HEADERS, start_col):
            cell = ws.cell(row=title_row, column=col_idx, value=header)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

        for field_path, metrics in analysis_result['field_metrics'].items():
            mapped_row = Config.FIELD_MAPPING.get(field_path)
            if mapped_row:
                Step3_ResultFiller._fill_field_metrics(ws, mapped_row, start_col, metrics, analysis_result, field_path)

        gene_row = None
        for field_path, metrics in analysis_result['field_metrics'].items():
            mapped_row = Config.FIELD_MAPPING.get(field_path)
            if mapped_row:
                if field_path == "Variants Include.Gene":
                    gene_row = mapped_row

        if gene_row is not None:
            time_col_index = Config.INDICATOR_HEADERS.index('7 Processing Time')
            time_col = start_col + time_col_index
            ws.cell(row=gene_row, column=time_col, value=qa_time)

        last_data_row = 4 + len(Config.FIELD_STRUCTURE)

        indicator_data = [
            analysis_result.get('std_total', 0),
            analysis_result.get('model_total', 0),
            analysis_result.get('correct_total', 0),
            analysis_result.get('false_assert_total', 0),
            'N/A',
            analysis_result.get('field_omissions_total', 0),
            'N/A',
            qa_time,
            'N/A'
        ]

        for col_idx, value in enumerate(indicator_data, start_col):
            ws.cell(row=last_data_row + 2, column=col_idx, value=value)

        wb.save(output_path)
        print(f"PMID {pmid} results appended to {output_path}")

    @staticmethod
    def _fill_field_metrics(ws, row, start_col, metrics, analysis_result, field_path):
        columns_mapping = [
            (6, 'std_count'),
            (7, 'model_count'),
            (8, 'correct'),
            (9, 'false_assert'),
            (10, 'logical_conflict'),
            (11, 'field_omissions'),
            (16, 'processing_time'),
        ]

        if field_path in Config.FIELD_GROUPS:
            group_name = Config.FIELD_GROUPS[field_path]
            group_fields = [f for f in Config.FIELD_GROUPS if Config.FIELD_GROUPS[f] == group_name]
            if field_path != group_fields[0]:
                return

            group_metrics = {
                'std_count': 0,
                'model_count': 0,
                'correct': 0,
                'false_assert': 0,
            }

            for f in group_fields:
                if f in analysis_result['field_metrics']:
                    fm = analysis_result['field_metrics'][f]
                    group_metrics['std_count'] = fm['std_count']
                    group_metrics['model_count'] = fm['model_count']
                    group_metrics['correct'] = fm['correct']
                    group_metrics['false_assert'] = fm['false_assert']
                    break

            for col_offset, key in columns_mapping:
                cell = ws.cell(row=row, column=start_col + col_offset - 6)
                if key == 'field_omissions' and group_name in ["Protein Change", "cDNA Change"]:
                    cell.value = group_metrics['std_count'] - group_metrics['correct']
                else:
                    cell.value = group_metrics.get(key, 0)
            return

        for col_offset, key in columns_mapping:
            cell = ws.cell(row=row, column=start_col + col_offset - 6)
            if key == 'field_omissions':
                if field_path in Config.OMISSION_FIELDS:
                    cell.value = metrics['std_count'] - metrics['correct']
                else:
                    cell.value = 0
            elif key == 'logical_conflict':
                cell.value = 'N/A'
            else:
                cell.value = metrics.get(key, 0)


class TimeExtractor:
    @staticmethod
    def extract(content):
        try:
            match = re.search(r'Processing Time \(seconds\):\s*(\d+\.\d+)', content)
            return float(match.group(1)) if match else None
        except:
            return None


class FileProcessor:
    @staticmethod
    def load_json(file_path, is_std_file=False):
        try:
            encodings = ['utf-8-sig', 'utf-8', 'gb18030', 'latin-1']
            content = None
            for enc in encodings:
                try:
                    with open(file_path, 'r', encoding=enc) as f:
                        content = f.read()
                        break
                except:
                    continue

            if not content:
                raise ValueError("All encoding attempts failed")

            def remove_json_comments(json_str):
                pattern = r'(\s*//.*?\n)|(/\*.*?\*/)'
                return re.sub(pattern, '', json_str, flags=re.DOTALL)

            def safe_json_parse(json_str):
                try:
                    return json.loads(json_str)
                except json.JSONDecodeError as e:
                    fixed = re.sub(r',\s*}(?=\s*})', '}', json_str)
                    fixed = re.sub(r'[\x00-\x1F\x7F]', '', fixed)
                    try:
                        return json.loads(fixed)
                    except:
                        import json5
                        return json5.loads(fixed)

            for pattern in [r'```json\s*(.*?)\s*```', r'\{.*\}', r'\[.*\]']:
                match = re.search(pattern, content, re.DOTALL)
                if match:
                    json_content = match.group(1) if pattern == r'```json\s*(.*?)\s*```' else match.group(0)
                    cleaned_content = remove_json_comments(json_content)
                    repaired_content = json_content  # No longer using complex repair logic
                    try:
                        parsed = safe_json_parse(repaired_content)
                        return parsed
                    except Exception as e:
                        continue

            return safe_json_parse(remove_json_comments(content))

        except Exception as e:
            raise Exception(f"File loading failed: {str(e)}")

    @staticmethod
    def extract_pmid(file_path):
        return os.path.splitext(os.path.basename(file_path))[0].split('_')[0]


if __name__ == "__main__":
    # Configuration Paths
    STD_DIR = r'./ref_json'  # Standard JSON files directory
    MODEL_DIR = r'../result_split'  # Model output files directory
    OUTPUT_PATH = r'../04result_xlsx_split'  # Output Excel file path
    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    # Initialize statistics variables
    total_files = 0
    success_pmids = []
    failed_pmids = {}

    print("="*50)
    print("Starting processing of publication extraction results comparison...")
    print("="*50)

    # Get list of standard files
    std_files = [f for f in os.listdir(STD_DIR) if f.endswith('.json')]
    total_files = len(std_files)
    print(f"Detected {total_files} publications to process")
    print()

    for std_file in std_files:
        # Extract PMID
        pmid = FileProcessor.extract_pmid(std_file)
        print(f"Processing PMID: {pmid}...")

        try:
            # Construct full path for standard file
            std_file_path = os.path.join(STD_DIR, std_file)

            # Initialize model file variables
            model_json = None
            model_file_used = None

            # Attempt to load model files
            model_files = []
            # Try matching new JSON file format (including model name and number)
            model_file_pattern = f'{pmid}_*.json'  # Match JSON files starting with PMID
            model_files = [os.path.join(MODEL_DIR, f) for f in os.listdir(MODEL_DIR) if fnmatch.fnmatch(f, model_file_pattern)]
            # Try matching JSON file
            model_file_json = f'{pmid}.json'
            model_file_json_path = os.path.join(MODEL_DIR, model_file_json)
            if os.path.exists(model_file_json_path):
                model_files.append(model_file_json_path)
            # Try matching TXT file
            model_file_txt_pattern = f'{pmid}_*.txt'
            model_files_txt = [os.path.join(MODEL_DIR, f) for f in os.listdir(MODEL_DIR) if fnmatch.fnmatch(f, model_file_txt_pattern)]
            model_files.extend(model_files_txt)

            # Check if any model files are found
            if not model_files:
                error_msg = "No matching model files found"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Attempt to load each matched model file
            for model_file in model_files:
                try:
                    model_json = FileProcessor.load_json(model_file)
                    model_file_used = model_file
                    print(f"✅ Successfully loaded model output file: {os.path.basename(model_file)}")
                    break
                except Exception as e:
                    print(f"\n⚠️ Model output file {os.path.basename(model_file)} loading failed: {str(e)}")
                    continue

            # If all model files fail to load, skip current iteration
            if not model_json:
                error_msg = "All model files failed to load"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Load standard file
            try:
                std_json = FileProcessor.load_json(std_file_path, is_std_file=True)
                print(f"✅ Successfully loaded standard file")
            except Exception as e:
                error_msg = f"Standard file loading failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Compare data
            try:
                comparator = Step1_DataComparator(std_json)
                analysis_result = comparator.compare(model_json)
                print(f"✅ Data comparison completed")
            except Exception as e:
                error_msg = f"Data comparison failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Extract processing time
            try:
                with open(model_file_used, 'r', encoding='utf-8') as f:
                    model_content = f.read()
                qa_time = TimeExtractor.extract(model_content)
                if qa_time:
                    print(f"⏱️ Processing time: {qa_time} seconds")
            except Exception as e:
                qa_time = None
                print(f"\n⚠️ Processing time extraction failed: {str(e)}")

            # Create/load template
            try:
                if os.path.exists(OUTPUT_PATH):
                    wb = load_workbook(OUTPUT_PATH)
                else:
                    wb = Step2_TemplateBuilder.create_template(OUTPUT_PATH)
            except Exception as e:
                error_msg = f"Template creation failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Fill results
            try:
                Step3_ResultFiller.fill_results(wb, OUTPUT_PATH, pmid, analysis_result, qa_time)
                print(f"💾 Results saved to Excel")
                success_pmids.append(pmid)
            except Exception as e:
                error_msg = f"Result filling failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg

            print(f"✅ PMID {pmid} processing completed")
            print("="*50)

        except Exception as e:
            error_msg = f"Unknown error: {str(e)}"
            print(f"\n⛔ Unknown error occurred while processing PMID {pmid}: {str(e)}")
            failed_pmids[pmid] = error_msg

    # Output statistics
    print("\n" + "="*50)
    print("📊 Processing Results Statistics")
    print("="*50)
    print(f"Total publications: {total_files}")
    print(f"Successfully processed: {len(success_pmids)} publications")
    print(f"Failed processing: {len(failed_pmids)} publications")
    print(f"Success rate: {len(success_pmids)/total_files*100:.1f}%")

    if failed_pmids:
        print("\n⛔ Failed PMIDs and Reasons:")
        for pmid, reason in failed_pmids.items():
            print(f"  {pmid}: {reason}")

    if success_pmids:
        print("\n✅ Successfully Processed PMIDs:")
        print(", ".join(success_pmids))
    else:
        print("\n⛔ No PMIDs successfully processed")

    print("\nAll file processing completed!")