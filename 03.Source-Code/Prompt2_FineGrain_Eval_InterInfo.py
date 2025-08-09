import fnmatch
import json
import re
import os
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension

# Fine grained evaluation of intermediate information extracted from LLMs

class Config:
    """Global Configuration Center"""
    # Comparison configuration
    EXACT_MATCH_FIELDS = {
        "Variants Include.Gene",
        "Variants Include.variants.HGVS",
        "Variants Include.variants.cDNA Change.transcript",
        "Variants Include.variants.cDNA Change.ref",
        "Variants Include.variants.cDNA Change.alt",
        "Variants Include.variants.cDNA Change.position",
        "Variants Include.variants.Protein Change.ref",
        "Variants Include.variants.Protein Change.alt",
        "Variants Include.variants.Protein Change.position",
        "Described Disease.Described Disease",
        "Described Disease.MONDO",
        "Experiment Method.Assay Method",
        "Experiment Method.Material used.Material Source",
        "Experiment Method.Material used.Material Name",
        "Experiment Method.Readout type",
        "Experiment Method.Readout description.Conclusion",
        "Experiment Method.Readout description.Molecular Effect",
        "Experiment Method.Biological replicates.Biological replicates",
        "Experiment Method.Technical replicates.Technical replicates",
        "Experiment Method.Basic positive control.Basic positive control",
        "Experiment Method.Basic negative control.Basic negative control",
        "Experiment Method.Approved assay.Approved assay"
    }

    SKIP_FIELDS = {
        "Variants Include.variants.Description in input context",
        "Experiment Method.Material used.Description",
        "Experiment Method.Readout description.Result Description",
        "Experiment Method.Biological replicates.Description",
        "Experiment Method.Technical replicates.Description",
        "Experiment Method.Basic positive control.Description",
        "Experiment Method.Basic negative control.Description",
        "Experiment Method.Validation controls P/LP",
        "Experiment Method.Validation controls B/LB",
        "Experiment Method.Statistical analysis method",
        "Experiment Method.Threshold for normal readout",
        "Experiment Method.Threshold for abnormal readout"
    }

    BOOLEAN_FIELDS = {
        "Experiment Method.Biological replicates.Biological replicates",
        "Experiment Method.Technical replicates.Technical replicates",
        "Experiment Method.Basic positive control.Basic positive control",
        "Experiment Method.Basic negative control.Basic negative control",
        "Experiment Method.Approved assay.Approved assay"
    }

    FIELD_GROUPS = {
        "Variants Include.variants.cDNA Change.ref": "cDNA Change",
        "Variants Include.variants.cDNA Change.alt": "cDNA Change",
        "Variants Include.variants.cDNA Change.position": "cDNA Change",
        "Variants Include.variants.Protein Change.ref": "Protein Change",
        "Variants Include.variants.Protein Change.alt": "Protein Change",
        "Variants Include.variants.Protein Change.position": "Protein Change",
    }

    AMINO_ACID_MAP = {
        # Three-letter/single-letter/full name mapping
        'ala': 'A', 'alanine': 'A', 'A': 'A',
        'arg': 'R', 'arginine': 'R', 'R': 'R',
        'asn': 'N', 'asparagine': 'N', 'N': 'N',
        'asp': 'D', 'aspartic acid': 'D', 'D': 'D',
        'cys': 'C', 'cysteine': 'C', 'C': 'C',
        'gln': 'Q', 'glutamine': 'Q', 'Q': 'Q',
        'glu': 'E', 'glutamic acid': 'E', 'E': 'E',
        'gly': 'G', 'glycine': 'G', 'G': 'G',
        'his': 'H', 'histidine': 'H', 'H': 'H',
        'ile': 'I', 'isoleucine': 'I', 'I': 'I',
        'leu': 'L', 'leucine': 'L', 'L': 'L',
        'lys': 'K', 'lysine': 'K', 'K': 'K',
        'met': 'M', 'methionine': 'M', 'M': 'M',
        'phe': 'F', 'phenylalanine': 'F', 'F': 'F',
        'pro': 'P', 'proline': 'P', 'P': 'P',
        'ser': 'S', 'serine': 'S', 'S': 'S',
        'thr': 'T', 'threonine': 'T', 'T': 'T',
        'trp': 'W', 'tryptophan': 'W', 'W': 'W',
        'tyr': 'Y', 'tyrosine': 'Y', 'Y': 'Y',
        'val': 'V', 'valine': 'V', 'V': 'V'
    }

    NUCLEIC_ACID_MAP = {
        'adenine': 'A', 'A': 'A',
        'thymine': 'T', 'T': 'T',
        'cytosine': 'C', 'C': 'C',
        'guanine': 'G', 'G': 'G',
        'uracil': 'U', 'U': 'U'
    }

    # Excel template configuration
    LEFT_COLUMNS = 5  # Number of fixed columns on the left
    HEADER_ROW = 2
    FIELD_START_ROW = 5
    PMID_COL = 6  # Starting column for literature metrics

    FIELD_STRUCTURE = [
        # Column 1, Column 2, Column 3, Column 4 (Comparison Requirement), Column 5 (Description)
        ["Variants Include", "Gene", "", "Number consistent with reference results", ""],
        ["", "variants", "HGVS", "Number consistent with reference results", ""],
        ["", "", "cDNA Change", "transcript", "Number consistent with reference results"],
        ["", "", "", "ref,alt,position", "Number consistent with reference results"],
        ["", "", "Protein Change", "ref,alt,position", "Number consistent with reference results"],
        ["", "", "Description in input context", "", "Length/number consistent with original text"],

        ["Described Disease", "Described Disease", "", "Consistent with reference results", ""],
        ["", "MONDO", "", "Consistent with reference results", ""],

        ["Experiment Method", "Assay Method", "", "Consistent with reference results", ""],
        ["", "Material used", "Material Source", "Consistent with reference results", ""],
        ["", "", "Material Name", "Consistent with reference results", ""],
        ["", "", "Description", "", "Length consistent with original text"],
        ["", "Readout type", "", "Consistent with reference results", ""],
        ["", "Readout description", "", "Length consistent with original text", ""],
        ["", "", "Variant", "", "Not compared"],
        ["", "", "Conclusion", "", "Consistent with reference results"],
        ["", "", "Molecular Effect", "", "Consistent with reference results"],
        ["", "", "Result Description", "", "Length consistent with original text"],
        ["", "Biological replicates", "Biological replicates", "yes/no", "Consistent with reference results"],
        ["", "", "Description", "", "Length consistent with original text"],
        ["", "Technical replicates", "Technical replicates", "yes/no", "Consistent with reference results"],
        ["", "", "Description", "", "Length consistent with original text"],
        ["", "Basic positive control", "Basic positive control", "yes/no", "Consistent with reference results"],
        ["", "", "Description", "", "Length consistent with original text"],
        ["", "Basic negative control", "Basic negative control", "yes/no", "Consistent with reference results"],
        ["", "", "Description", "", "Length consistent with original text"],
        ["", "Validation controls P/LP", "Validation controls P/LP", "yes/no", "Not compared"],
        ["", "", "Counts", "", "Not compared"],
        ["", "Validation controls B/LB", "Validation controls B/LB", "yes/no", "Not compared"],
        ["", "", "Counts", "", "Not compared"],
        ["", "Statistical analysis method", "", "Not compared", ""],
        ["", "Threshold for normal readout", "Source", "Not compared", ""],
        ["", "Threshold for abnormal readout", "Source", "Not compared", ""],
        ["", "Approved assay", "Approved assay", "yes/no", "Consistent with reference results"]
    ]

    INDICATOR_HEADERS = [
        '1 Standard Total', '1 Total Output', '1 Correct Count', '4 False Assertions', '5 Logical Contradictions',
        '2 Field Omissions', '3 Standard Yes', '3 Yes Correct', '3 Standard No', '3 No Correct',
        '6 Consistency Rate', '7 Processing Time', '9 Final Rating Consistency'
    ]

    FIELD_MAPPING= {
        # Variants Include section (rows 4-9)
        "Variants Include.Gene": 4,
        "Variants Include.variants.HGVS": 5,
        "Variants Include.variants.cDNA Change.transcript": 6,
        "Variants Include.variants.cDNA Change.ref": 7,
        "Variants Include.variants.cDNA Change.alt": 7,
        "Variants Include.variants.cDNA Change.position": 7,
        "Variants Include.variants.Protein Change.ref": 8,
        "Variants Include.variants.Protein Change.alt": 8,
        "Variants Include.variants.Protein Change.position": 8,
        "Variants Include.variants.Description in input context": 9,

        # Described Disease section (rows 10-11)
        "Described Disease.Described Disease": 10,
        "Described Disease.MONDO": 11,

        # Experiment Method section (rows 12-36)
        "Experiment Method.Assay Method": 12,
        "Experiment Method.Material used.Material Source": 13,
        "Experiment Method.Material used.Material Name": 14,
        "Experiment Method.Material used.Description": 15,
        "Experiment Method.Readout type": 16,
        "Experiment Method.Readout description.Result Description": 21,
        "Experiment Method.Readout description.Conclusion": 19,
        "Experiment Method.Readout description.Molecular Effect": 20,
        "Experiment Method.Biological replicates.Biological replicates": 22,
        "Experiment Method.Biological replicates.Description": 23,
        "Experiment Method.Technical replicates.Technical replicates": 24,
        "Experiment Method.Technical replicates.Description": 25,
        "Experiment Method.Basic positive control.Basic positive control": 26,
        "Experiment Method.Basic positive control.Description": 27,
        "Experiment Method.Basic negative control.Basic negative control": 28,
        "Experiment Method.Basic negative control.Description": 29,
        "Experiment Method.Approved assay.Approved assay": 37
    }

    # Fields for omission statistics
    OMISSION_FIELDS = {
        "Variants Include.variants.Protein Change.ref",
        "Variants Include.variants.Protein Change.alt",
        "Variants Include.variants.Protein Change.position",
        "Experiment Method.Assay Method"
    }


class Step1_DataComparator:
    def _normalize_amino_acid(self, value):
        """Enhanced amino acid normalization (handles deletions)"""
        value = str(value).strip().lower()
        # Handle deletion marker
        if "deletion" in value:
            return ""
        # Handle bracketed format (e.g., Leucine (L))
        if '(' in value and ')' in value:
            bracket_part = value.split('(')[-1].split(')')[0].strip()
            if len(bracket_part) == 1:
                return bracket_part.upper()
        # Handle full name/three-letter/single-letter forms
        for key in Config.AMINO_ACID_MAP:
            if key in value:
                return Config.AMINO_ACID_MAP[key]
        return value[0].upper() if value else ""

    def _normalize_nucleic_acid(self, value, is_dna=True):
        """Normalize nucleic acid representation to single letter"""
        value = re.sub(r'[^a-zA-Z]', '', str(value)).upper()
        base_map = Config.NUCLEIC_ACID_MAP
        # Filter U for DNA, T for RNA
        if is_dna:
            return base_map.get(value, '').replace('U', '')
        else:
            return base_map.get(value, '').replace('T', '')

    def _normalize_position(self, value):
        """Extract pure numeric position"""
        return ''.join(re.findall(r'\d+', str(value)))

    """Step 1: Data Comparator"""
    def __init__(self, std_json):
        self.std_json = std_json
        self.field_metrics = defaultdict(lambda: {
            'std_count': 0,        # 1 Standard Total
            'model_count': 0,      # 1 Total Output
            'correct': 0,         # 1 Correct Count
            'false_assert': 0,    # 4 False Assertions
            'std_yes': 0,         # 3 Standard Yes
            'std_no': 0,          # 3 Standard No
            'correct_yes': 0,     # 3 Yes Correct
            'correct_no': 0,     # 3 No Correct
            'field_omissions': 0,  # Field omission counter
            'correct_details': [],  #
            'std_values': [],  # Store all standard field values
            'model_values': [],  # Store all model output values
            'matched_std_values': set()  # Store matched standard values
        })
        self.qa_time = None       # 7 Processing Time
        self.field_details = []   # Detailed comparison records

    def compare(self, model_json):
        """Perform deep comparison"""
        # Standard side traversal: use standard data as current_node, model data as compare_node
        self._traverse_node(self.std_json, model_json, is_std=True)
        # Model side traversal: use model data as current_node, standard data as compare_node
        self._traverse_node(model_json, self.std_json, is_std=False)
        self._process_field_groups()
        return self._calculate_metrics()

    def _process_field_groups(self):
        """Optimized field group processing logic"""
        group_metrics = defaultdict(lambda: {
            'std_variants': set(),
            'model_variants': set(),
            'std_count': 0,
            'model_count': 0,
            'correct': 0,
            'false_assert': 0
        })

        ND_MARKER = "N.D."
        DELETION_MARKER = "[DEL]"

        # === Standard variant processing optimization ===
        for group_name in {"cDNA Change", "Protein Change"}:
            group_fields = [f for f in Config.FIELD_GROUPS if Config.FIELD_GROUPS[f] == group_name]
            ref_field, alt_field, pos_field = group_fields[0], group_fields[1], group_fields[2]

            # Get standard values (using safe indexing)
            std_refs = self.field_metrics[ref_field]['std_values']
            std_alts = self.field_metrics[alt_field]['std_values']
            std_positions = self.field_metrics[pos_field]['std_values']

            # Calculate maximum valid length
            max_len = max(len(std_refs), len(std_alts), len(std_positions))
            valid_std = []

            for i in range(max_len):
                # Safely get field values
                ref = std_refs[i] if i < len(std_refs) else ""
                alt = std_alts[i] if i < len(std_alts) else ""
                pos = std_positions[i] if i < len(std_positions) else ""

                # Normalize processing (preserve empty strings)
                ref = self._normalize_component(ref_field, i, group_name)
                alt = self._normalize_component(alt_field, i, group_name)
                pos = self._normalize_component(pos_field, i, group_name)

                # Optimized filtering logic: filter only when all fields are invalid
                if all(v in (ND_MARKER, "") for v in (ref, alt, pos)):
                    continue
                valid_std.append((ref, alt, pos))

            group_metrics[group_name]['std_variants'] = set(valid_std)
            group_metrics[group_name]['std_count'] = len(valid_std)

        # === Model variant processing optimization ===
        for group_name in group_metrics:
            group_fields = [f for f in Config.FIELD_GROUPS if Config.FIELD_GROUPS[f] == group_name]
            ref_field, alt_field, pos_field = group_fields[0], group_fields[1], group_fields[2]

            model_refs = self.field_metrics[ref_field]['model_values']
            model_alts = self.field_metrics[alt_field]['model_values']
            model_positions = self.field_metrics[pos_field]['model_values']

            # Use same logic to process model output
            max_len = max(len(model_refs), len(model_alts), len(model_positions))
            valid_model = []

            for i in range(max_len):
                ref = model_refs[i] if i < len(model_refs) else ""
                alt = model_alts[i] if i < len(model_alts) else ""
                pos = model_positions[i] if i < len(model_positions) else ""

                ref = self._normalize_component(ref_field, i, group_name, is_model=True)
                alt = self._normalize_component(alt_field, i, group_name, is_model=True)
                pos = self._normalize_component(pos_field, i, group_name, is_model=True)

                # Special handling for Protein deletion
                if group_name == "Protein Change" and "deletion" in str(alt).lower():
                    alt = DELETION_MARKER

                valid_model.append((ref, alt, pos))

            group_metrics[group_name]['model_variants'] = set(valid_model)
            group_metrics[group_name]['model_count'] = len(valid_model)

        # Add debug logging in model variant processing
        for i in range(max_len):
            ref = model_refs[i] if i < len(model_refs) else ""
            print(f"[DEBUG] Model value - {ref_field}[{i}]: {ref}")

        # === Comparison logic optimization ===
        for group_name in group_metrics:
            std_set = group_metrics[group_name]['std_variants']
            model_set = group_metrics[group_name]['model_variants']

            # Correct matches: allow partial field emptiness but position matching
            correct = sum(1 for m in model_set if m in std_set)
            # False assertions: model-specific assertions without empty values
            false_assert = len([m for m in model_set if m not in std_set and all(v != "" for v in m)])

            group_metrics[group_name]['correct'] = correct
            group_metrics[group_name]['false_assert'] = false_assert

        # === Enhanced debug logging ===
        for field_path in Config.FIELD_GROUPS:
            group_name = Config.FIELD_GROUPS[field_path]
            self.field_metrics[field_path].update({
                'std_count': group_metrics[group_name]['std_count'],
                'model_count': group_metrics[group_name]['model_count'],
                'correct': group_metrics[group_name]['correct'],
                'false_assert': group_metrics[group_name]['false_assert']
            })
            print(f"[DEBUG] {group_name} group statistics: std={group_metrics[group_name]['std_count']}, "
                  f"model={group_metrics[group_name]['model_count']}, "
                  f"correct={group_metrics[group_name]['correct']}")

    def _normalize_component(self, field_path, index, group_name, is_model=False):
        """Enhanced normalization logic"""
        source = 'model_values' if is_model else 'std_values'
        values = self.field_metrics[field_path][source]
        value = values[index] if index < len(values) else ""

        # Preserve original value for debugging
        original = str(value).strip()
        normalized = original

        # Handle special markers
        if original.upper() == 'N.D.':
            return "N.D."

        # Apply normalization rules
        try:
            if "Protein" in group_name:
                if "ref" in field_path or "alt" in field_path:
                    normalized = self._normalize_amino_acid(value)
                else:
                    normalized = self._normalize_position(value)
            else:
                if "ref" in field_path or "alt" in field_path:
                    normalized = self._normalize_nucleic_acid(value)
                else:
                    normalized = self._normalize_position(value)
        except Exception as e:
            print(f"[WARN] Normalization failed: {field_path}[{index}] '{original}' -> {str(e)}")
            normalized = original

        print(f"[DEBUG] Normalization: {field_path}[{index}] '{original}' -> '{normalized}'")
        return normalized

    def _traverse_node(self, current_node, compare_node, is_std, path=[]):
        current_path = '.'.join(path)
        print(f"[DEBUG] Traversing path: {current_path}, is_std={is_std}, type: {type(current_node)}")
        """Optimized traversal logic ensuring complete recursive traversal"""
        if isinstance(current_node, dict):
            for key in current_node:
                new_path = path + [key]
                # Ensure comparison node's structure is correct
                compare_value = compare_node.get(key, None) if isinstance(compare_node, dict) else None
                self._traverse_node(current_node[key], compare_value, is_std, new_path)
        elif isinstance(current_node, list):
            # Traverse list ensuring all elements are processed
            for idx, item in enumerate(current_node):
                compare_item = compare_node[idx] if (isinstance(compare_node, list) and idx < len(compare_node)) else None
                self._traverse_node(item, compare_item, is_std, path.copy())
        else:
            # Ensure correct parameter order when processing leaf nodes
            if is_std:
                # Standard side: std_val=current_node, model_val=compare_node
                self._process_leaf_node(current_node, compare_node, is_std, path)
            else:
                # Model side: std_val=compare_node (standard value), model_val=current_node (model value)
                self._process_leaf_node(compare_node, current_node, is_std, path)

    def _process_node(self, current_node, compare_node, is_std, path, key):
        """Process dictionary node (fixed version)"""
        new_path = path + [key]

        # Add type safety check
        if not isinstance(compare_node, dict):
            if is_std:
                self._handle_missing_field('.'.join(new_path))
            return

        if key in compare_node:
            self._traverse_node(
                current_node[key],
                compare_node[key],
                is_std,
                new_path
            )
        elif is_std:
            self._handle_missing_field('.'.join(new_path))

    def _process_list(self, item, compare_node, is_std, path, idx):
        """Process list node (fixed version)"""
        # Add list type check
        if not isinstance(compare_node, list):
            compare_item = None
        else:
            try:
                compare_item = compare_node[idx] if idx < len(compare_node) else None
            except:
                compare_item = None

        self._traverse_node(
            item,
            compare_item,
            is_std,
            path.copy()
        )

    def _handle_missing_field(self, field_path):
        """Fixed missing field handling"""
        if field_path in self.field_metrics:
            # Increment field omission counter
            self.field_metrics[field_path]['field_omissions'] += 1
            # Update standard count
            self.field_metrics[field_path]['std_count'] += 1

    def _process_leaf_node(self, std_val, model_val, is_std, path):
        """Optimized leaf node processing logic, preserving original structure with new improvements"""
        field_path = '.'.join(path)
        print(f"[DEBUG] Processing field: {field_path}, is_std={is_std}")
        print(f"[DEBUG] Standard value: {std_val}, Model value: {model_val}")
        metrics = self.field_metrics[field_path]

        # Enhanced N.D. recognition logic (fix)
        if not is_std:
            model_str = str(model_val).strip().upper() if model_val is not None else ""
            # Use regex to match various N.D. formats
            if re.match(r'^N\.?D\.?$', model_str, re.IGNORECASE):
                print(f"[DEBUG] Detected N.D. value: {model_str}")
                metrics['model_values'].append("N.D.")
                metrics['model_count'] += 1
                return  # Skip subsequent metric calculations

        # Preserve SKIP_FIELDS handling logic
        if field_path in Config.SKIP_FIELDS:
            if is_std:
                metrics['std_values'].append(str(std_val))
                metrics['std_count'] += 1
            else:
                model_str = str(model_val) if model_val is not None else ""
                metrics['model_values'].append(model_str)
                metrics['model_count'] += 1

                std_content = str(std_val).lower()
                model_content = model_str.lower()
                if std_content == model_content:
                    metrics['correct'] += 1
                    detail = f"Standard: {std_val}, Output: {model_val}"
                    metrics['correct_details'].append(detail)
            return

        # Preserve boolean field handling logic
        if field_path in Config.BOOLEAN_FIELDS:
            self._process_boolean(std_val, model_val, is_std, field_path)
            return

        # Handle model side with correct model_val normalization
        # Handle EXACT_MATCH_FIELDS (add mutual inclusion check)
        if field_path in Config.EXACT_MATCH_FIELDS:
            if is_std:
                # Standard side processing remains unchanged
                clean_val = re.sub(r'[^a-zA-Z0-9]', '', str(std_val)).lower()
                metrics['std_values'].append(clean_val)
                metrics['std_count'] += 1
            else:
                # Model side processing with inclusion logic
                model_clean = re.sub(r'[^a-zA-Z0-9]', '', str(model_val)).lower() if model_val else ''
                metrics['model_values'].append(model_clean)
                metrics['model_count'] += 1

                # Determine field type
                is_boolean = field_path in Config.BOOLEAN_FIELDS
                is_group = field_path in Config.FIELD_GROUPS

                # Matching logic
                if is_boolean or is_group:
                    matched = model_clean in metrics['std_values']
                else:
                    # Add continuous inclusion check
                    matched = any(
                        std in model_clean or model_clean in std
                        for std in metrics['std_values']
                    )

                if matched:
                    metrics['correct'] += 1
                else:
                    metrics['false_assert'] += 1

            return

    def _is_contained(self, a, b):
        """Check for continuous inclusion"""
        return a in b or b in a

    def _process_boolean(self, std_val, model_val, is_std, field_path):
        """Optimized boolean field processing logic"""
        metrics = self.field_metrics[field_path]
        if is_std:
            # Standard side statistics
            metrics['std_count'] += 1
            std_val = str(std_val).lower()
            if std_val == 'yes':
                metrics['std_yes'] += 1
            elif std_val == 'no':
                metrics['std_no'] += 1
        else:
            # Model side statistics
            metrics['model_count'] += 1
            std_val = str(std_val).lower()
            model_val = str(model_val).lower() if model_val else ''

            # Only evaluate when standard has a clear yes/no value
            if std_val == 'yes':
                if model_val == 'yes':
                    metrics['correct_yes'] += 1
                    metrics['correct'] += 1
                else:
                    metrics['false_assert'] += 1
            elif std_val == 'no':
                if model_val == 'no':
                    metrics['correct_no'] += 1
                    metrics['correct'] += 1
                else:
                    metrics['false_assert'] += 1
            else:
                # Do not count correctness if standard lacks clear yes/no
                pass

    def _process_exact_match(self, std_val, model_val, is_std, field_path):
        """Exact match field processing"""
        metrics = self.field_metrics[field_path]
        if is_std:
            metrics['std_count'] += 1
        else:
            metrics['model_count'] += 1
            std_clean = re.sub(r'\W+', '', str(std_val)).lower()
            model_clean = re.sub(r'\W+', '', str(model_val)).lower() if model_val else ''
            if std_clean == model_clean:
                metrics['correct'] += 1
            else:
                metrics['false_assert'] += 1

    def _calculate_metrics(self):
        """Optimized metrics calculation logic"""
        total = defaultdict(int)
        processed_groups = set()

        # First pass: handle field groups and regular fields
        for field_path, metrics in self.field_metrics.items():
            # Handle field group logic
            if field_path in Config.FIELD_GROUPS:
                group_name = Config.FIELD_GROUPS[field_path]
                if group_name in processed_groups:
                    continue

                # Get metrics for the first field in the group
                group_fields = [f for f in Config.FIELD_GROUPS if Config.FIELD_GROUPS[f] == group_name]
                main_field = group_fields[0]
                main_metrics = self.field_metrics[main_field]

                # Accumulate group metrics (count each group once)
                total['std_total'] += main_metrics['std_count']
                total['model_total'] += main_metrics['model_count']
                total['correct_total'] += main_metrics['correct']
                total['false_assert_total'] += main_metrics['false_assert']

                processed_groups.add(group_name)
                continue

            # Handle regular fields (non-field groups)
            if field_path not in Config.FIELD_GROUPS.values():
                total['std_total'] += metrics['std_count']
                total['model_total'] += metrics['model_count']
                total['correct_total'] += metrics['correct']
                total['false_assert_total'] += metrics['false_assert']
                # Add non-field group field omission calculation
                total['field_omissions_total'] += max(
                    0, metrics['std_count'] - metrics['model_count']
                )

            # Handle boolean fields
            if field_path in Config.BOOLEAN_FIELDS:
                total['std_yes_total'] += metrics['std_yes']
                total['correct_yes_total'] += metrics['correct_yes']
                total['std_no_total'] += metrics['std_no']
                total['correct_no_total'] += metrics['correct_no']

            # Handle field omissions (only for specified fields)
            if field_path in Config.OMISSION_FIELDS:
                valid_model_count = metrics['model_count'] - metrics['model_values'].count("N.D.")
                total['field_omissions_total'] += max(
                    0,
                    metrics['std_count'] - metrics['correct'] -
                    (metrics['model_count'] - valid_model_count)
                )

        # Return calculated results
        return {
            'std_total': total['std_total'],
            'model_total': total['model_total'],
            'correct_total': total['correct_total'],
            'false_assert_total': total['false_assert_total'],
            'field_omissions_total': total['field_omissions_total'],
            'std_yes_total': total['std_yes_total'],
            'correct_yes_total': total['correct_yes_total'],
            'std_no_total': total['std_no_total'],
            'correct_no_total': total['correct_no_total'],
            'omission_fields_std_total': sum(
                self.field_metrics[f]['std_count']
                for f in Config.OMISSION_FIELDS
            ),
            'field_metrics': self.field_metrics,
            'qa_time': self.qa_time,
            'field_details': self.field_details
        }


class Step2_TemplateBuilder:
    """Step 2: Excel Template Builder"""
    @staticmethod
    def create_template(output_path):
        """Create/load template"""
        if os.path.exists(output_path):
            return load_workbook(output_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Results"

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 8
        for col in range(6, 50):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Add fixed headers
        # Merge cells
        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:D1')
        ws.merge_cells('E1:E3')

        # First row content
        ws['A1'] = "Model Name:"
        ws['C1'] = "Deployment Method"
        ws['E1'] = "Metrics"

        # Second and third row content
        ws.merge_cells('A2:D3')
        ws['A2'] = "1 Information Extraction Accuracy 2 Information Completeness 3 Classification Performance 4 Generation Quality (Hallucination) 5 Interpretability 6 Consistency 7 Efficiency 8 Scalability 9 Clinical Utility"

        # Set styles
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for cell in ws['A1:E3']:
            for c in cell:
                c.border = thin_border
                c.alignment = Alignment(horizontal='left', vertical='center')

        for cell in ws['A4:E' + str(len(Config.FIELD_STRUCTURE)+4)]:
            for c in cell:
                c.border = thin_border
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Add field structure
        field_start_row = 4
        for idx, row in enumerate(Config.FIELD_STRUCTURE, start=field_start_row):
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Add summary row
        ws.append([''] * 5 + ['Per Column', 'Summary'])

        # Freeze panes
        ws.freeze_panes = "F4"

        # Save template
        wb.save(output_path)
        return wb


class Step3_ResultFiller:
    @staticmethod
    def _find_next_column(ws):
        """Find the next available column"""
        return ws.max_column + 1

    """Step 3: Result Filler"""
    @staticmethod
    def fill_results(wb, output_path, pmid, analysis_result, qa_time):
        """Integrated optimized filling logic"""
        ws = wb.active

        # Find starting column
        start_col = ws.max_column + 1 if ws.max_column > Config.LEFT_COLUMNS else Config.PMID_COL
        start_row = Config.HEADER_ROW

        # Fill PMID identifier
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=start_row, end_column=start_col+len(Config.INDICATOR_HEADERS)-1)
        pmid_cell = ws.cell(row=start_row, column=start_col, value=f"PMID_{pmid}")
        pmid_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Fill metric headers
        title_row = start_row + 1
        for col_idx, header in enumerate(Config.INDICATOR_HEADERS, start_col):
            cell = ws.cell(row=title_row, column=col_idx, value=header)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

        # Fill field-level metrics
        for field_path, metrics in analysis_result['field_metrics'].items():
            mapped_row = Config.FIELD_MAPPING.get(field_path)
            if mapped_row:
                Step3_ResultFiller._fill_field_metrics(
                    ws,
                    mapped_row,
                    start_col,
                    metrics,
                    analysis_result,
                    field_path
                )

        # Add processing time record for gene row
        gene_row = None
        for field_path, metrics in analysis_result['field_metrics'].items():
            mapped_row = Config.FIELD_MAPPING.get(field_path)
            if mapped_row:
                if field_path == "Variants Include.Gene":
                    gene_row = mapped_row

        # Record processing time in gene row
        if gene_row is not None:
            time_col_index = Config.INDICATOR_HEADERS.index('7 Processing Time')
            time_col = start_col + time_col_index
            ws.cell(row=gene_row, column=time_col, value=qa_time)

        # Calculate last data row
        last_data_row = 4 + len(Config.FIELD_STRUCTURE)

        # Fill summary metrics
        indicator_data = [
            analysis_result.get('std_total', 0),
            analysis_result.get('model_total', 0),
            analysis_result.get('correct_total', 0),
            analysis_result.get('false_assert_total', 0),
            'N/A',  # 5 Logical Contradictions
            analysis_result.get('field_omissions_total', 0),
            analysis_result.get('std_yes_total', 0),
            analysis_result.get('correct_yes_total', 0),
            analysis_result.get('std_no_total', 0),
            analysis_result.get('correct_no_total', 0),
            'N/A',  # 6 Consistency Rate
            qa_time,
            'N/A'  # 9 Final Rating Consistency
        ]

        # Write indicator data
        for col_idx, value in enumerate(indicator_data, start_col):
            ws.cell(row=last_data_row + 2, column=col_idx, value=value)

        # Fill field details
        detail_row = start_row + 3
        for detail in analysis_result.get('field_details', []):
            if any(skip in detail['field_path'] for skip in Config.SKIP_FIELDS):
                continue

            # Fill field comparison row
            ws.cell(row=detail_row, column=1, value=detail['field_path'])
            ws.cell(row=detail_row, column=start_col, value=detail['std_val'])
            ws.cell(row=detail_row, column=start_col+1, value=detail['model_val'])
            ws.cell(row=detail_row, column=start_col+2, value='yes' if detail['is_correct'] else 'no')

            detail_row += 1

        # Add additional record for standard total sum in field omissions column
        try:
            omission_col_idx = Config.INDICATOR_HEADERS.index('2 Field Omissions')
            omission_col = start_col + omission_col_idx
        except ValueError:
            omission_col = None

        if omission_col:
            extra_info_row = last_data_row + 3
            ws.cell(row=extra_info_row, column=omission_col,
                    value=f"{analysis_result.get('omission_fields_std_total', 0)}")
            cell = ws.cell(row=extra_info_row, column=omission_col)
            cell.font = Font(italic=True, size=9)
            cell.alignment = Alignment(horizontal='right', vertical='center')

        # Set styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=start_row, max_row=detail_row, min_col=start_col):
            for cell in row:
                cell.border = thin_border

        # Save workbook
        wb.save(output_path)
        print(f"PMID {pmid} results appended to {output_path}")

    @staticmethod
    def _fill_field_metrics(ws, row, start_col, metrics, analysis_result, field_path):
        """Enhanced filling logic maintaining original header structure"""
        columns_mapping = [
            (6, 'std_count'),        # 1 Standard Total
            (7, 'model_count'),      # 1 Total Output
            (8, 'correct'),          # 1 Correct Count
            (9, 'false_assert'),     # 4 False Assertions
            (10, 'logical_conflict'),# 5 Logical Contradictions (placeholder)
            (11, 'field_omissions'), # 2 Field Omissions
            (12, 'std_yes'),         # 3 Standard Yes
            (13, 'correct_yes'),     # 3 Yes Correct
            (14, 'std_no'),          # 3 Standard No
            (15, 'correct_no'),      # 3 No Correct
            (16, 'consistency_rate'),# 6 Consistency Rate (placeholder)
            (17, 'processing_time'), # 7 Processing Time
            (18, 'final_rating')     # 9 Final Rating Consistency (placeholder)
        ]

        # Handle field group logic
        group_fields = []
        if field_path in Config.FIELD_GROUPS:
            group_name = Config.FIELD_GROUPS[field_path]
            group_fields = [f for f in Config.FIELD_GROUPS
                            if Config.FIELD_GROUPS[f] == group_name]
            if field_path != group_fields[0]:
                return

        # Handle descriptive fields (SKIP_FIELDS)
        if field_path in Config.SKIP_FIELDS:
            target_col = start_col + 2

            ws.cell(row=row, column=start_col+0, value=metrics['std_count'])
            ws.cell(row=row, column=start_col+1, value=metrics['model_count'])
            return

        # Process field group logic
        if field_path in Config.FIELD_GROUPS:
            group_name = Config.FIELD_GROUPS[field_path]
            all_fields_in_group = [f for f, g in Config.FIELD_GROUPS.items() if g == group_name]
            if field_path != all_fields_in_group[0]:
                return

            group_metrics = {
                'std_count': 0,
                'model_count': 0,
                'correct': 0,
                'false_assert': 0,
            }

            for f in all_fields_in_group:
                if f in analysis_result['field_metrics']:
                    fm = analysis_result['field_metrics'][f]
                    group_metrics['std_count'] = fm['std_count']
                    group_metrics['model_count'] = fm['model_count']
                    group_metrics['correct'] = fm['correct']
                    group_metrics['false_assert'] = fm['false_assert']
                    break

            for col_offset, key in columns_mapping:
                cell = ws.cell(row=row, column=start_col + col_offset - 6)
                if key in group_metrics:
                    cell.value = group_metrics[key]
                elif key == 'field_omissions' and group_name in ["Protein Change", "cDNA Change"]:
                    cell.value = group_metrics['std_count'] - group_metrics['correct']
                else:
                    cell.value = metrics.get(key, 0)

            return

        # Maintain original column order filling logic
        for col_offset, key in columns_mapping:
            cell = ws.cell(row=row, column=start_col + col_offset - 6)
            if field_path in Config.SKIP_FIELDS and key == 'correct':
                cell.value = "；".join(metrics.get('correct_details', []))
                continue
            if key == 'std_count':
                cell.value = metrics.get('std_count', 0)
            elif key == 'model_count':
                cell.value = metrics.get('model_count', 0)
            elif key == 'logical_conflict':
                cell.value = 'N/A'
            elif key == 'consistency_rate':
                cell.value = ''
            elif key == 'final_rating':
                cell.value = 'N/A'
            else:
                cell.value = metrics.get(key, 0)

        # Handle boolean fields
        if field_path in Config.BOOLEAN_FIELDS:
            ws.cell(row=row, column=start_col+12-6, value=metrics.get('std_yes', 0))
            ws.cell(row=row, column=start_col+13-6, value=metrics.get('correct_yes', 0))
            ws.cell(row=row, column=start_col+14-6, value=metrics.get('std_no', 0))
            ws.cell(row=row, column=start_col+15-6, value=metrics.get('correct_no', 0))

    def _fill_special_metrics(ws, row, start_col, metrics, field_path):
        """Handle metrics requiring special calculations"""
        cell = ws.cell(row=row, column=start_col+11-6)
        cell.value = metrics.get('std_count', 0) - metrics.get('correct', 0)

    @staticmethod
    def _log_intermediate_results(analysis_result):
        """Enhanced logging output (display all metrics)"""
        print("\n=== Field-level Statistics Details ===")
        for field, metrics in analysis_result['field_metrics'].items():
            print(f"\nField Path: {field}")
            print(f"├─ 1 Standard Total: {metrics['std_count']}")
            print(f"├─ 1 Total Output: {metrics['model_count']}")
            print(f"├─ 1 Correct Count: {metrics['correct']}")
            print(f"├─ 4 False Assertions: {metrics['false_assert']}")
            print(f"├─ 5 Logical Contradictions: {metrics.get('logical_conflict', 0)}")
            print(f"├─ 2 Field Omissions: {metrics['std_count'] - metrics['correct']}")

            if field in Config.BOOLEAN_FIELDS:
                print(f"├─ 3 Standard Yes: {metrics['std_yes']}")
                print(f"├─ 3 Yes Correct: {metrics['correct_yes']}")
                print(f"├─ 3 Standard No: {metrics['std_no']}")
                print(f"└─ 3 No Correct: {metrics['correct_no']}")
            else:
                print(f"├─ 6 Consistency Rate: {metrics['correct']/metrics['model_count']:.2%}" if metrics['model_count'] else "├─ 6 Consistency Rate: N/A")
                print(f"└─ 9 Final Rating: {metrics.get('final_rating', 'N/A')}")


class TimeExtractor:
    """Time Extractor"""
    @staticmethod
    def extract(content):
        """Enhanced time extraction"""
        try:
            match = re.search(r'Q&A Time \(seconds\):\s*(\d+\.\d+)', content)
            return float(match.group(1)) if match else None
        except:
            return None


class FileProcessor:
    """File Processing Center"""
    @staticmethod
    def fix_illegal_escapes(json_str):
        """
        Fix illegal escape sequences in JSON string
        Replace single backslash followed by an invalid escape character with double backslash
        Valid escape characters include: " \ / b f n r t u (followed by 4 hex digits)
        """
        illegal_escape_pattern = r'\\([^"\\/bfnrtu])'
        return re.sub(illegal_escape_pattern, r'\\\\\1', json_str)

    @staticmethod
    def structural_repair(json_str):
        """Intelligent structure repair"""
        # Step 1: Fix illegal escape characters
        json_str = FileProcessor.fix_illegal_escapes(json_str)

        # Step 2: Balance brackets and quotes
        stack = []
        in_string = False
        result = []
        i = 0

        # First pass: Balance brackets and quotes
        while i < len(json_str):
            char = json_str[i]

            # Handle escape characters
            if char == '\\' and i+1 < len(json_str):
                result.append(char + json_str[i+1])
                i += 2
                continue

            if char == '"':
                in_string = not in_string

            if not in_string:
                if char in ('{', '['):
                    stack.append(char)
                elif char == '}' and stack and stack[-1] == '{':
                    stack.pop()
                elif char == ']' and stack and stack[-1] == '[':
                    stack.pop()

            result.append(char)
            i += 1

        # Complete unclosed brackets
        for left in reversed(stack):
            result.append('}' if left == '{' else ']')

        # Second pass: Fix string truncation
        repaired = []
        in_string = False
        quote_count = 0
        for char in result:
            if char == '"' and (len(repaired) == 0 or repaired[-1] != '\\'):
                quote_count += 1
                in_string = not in_string
            repaired.append(char)

        # Complete odd number of quotes
        if quote_count % 2 != 0:
            repaired.append('"')

        # Third pass: Fix trailing truncation
        repaired_str = ''.join(repaired)
        last_char = repaired_str[-1] if repaired_str else ''
        if not re.search(r'[\]}\d"truefalsenull]$', last_char):
            if re.search(r':\s*[{\[]', repaired_str):
                repaired_str += ']}' if '{' in repaired_str else ']]'
            else:
                repaired_str += '}' if '{' in repaired_str else ']'

        return repaired_str

    @staticmethod
    def load_json(file_path, is_std_file=False):
        """Intelligent JSON loading (enhanced debug version)"""
        try:
            print(f"\n[DEBUG] {'Loading standard file' if is_std_file else 'Loading model output file'}: {file_path}")

            # Try multiple encoding methods
            encodings = ['utf-8-sig', 'utf-8', 'gb18030', 'latin-1']
            content = None
            for enc in encodings:
                try:
                    with open(file_path, 'r', encoding=enc) as f:
                        content = f.read()
                        print(f"[DEBUG] Successfully read file using encoding {enc}")
                        break
                except:
                    continue

            if not content:
                raise ValueError("All encoding attempts failed")

            # Print file header information
            print(f"[DEBUG] File content preview (first 500 characters):\n{content[:500]}...")
            print(f"[DEBUG] File total length: {len(content)} characters")

            # Remove JSON comments
            def remove_json_comments(json_str):
                """Remove comments from JSON using regex"""
                pattern = r'(\s*//.*?\n)|(/\*.*?\*/)'
                return re.sub(pattern, '', json_str, flags=re.DOTALL)

            # Tolerant JSON parsing
            def safe_json_parse(json_str):
                """JSON parsing with automatic repair"""
                try:
                    return json.loads(json_str)
                except json.JSONDecodeError as e:
                    fixed = re.sub(r',\s*}(?=\s*})', '}', json_str)
                    fixed = re.sub(r'[\x00-\x1F\x7F]', '', fixed)
                    try:
                        return json.loads(fixed)
                    except:
                        import json5
                        return json5.loads(fixed)

            # Try multiple JSON extraction patterns
            for pattern_idx, pattern in enumerate([
                r'```json\s*(.*?)\s*```',  # Markdown JSON code block
                r'\{.*\}',                 # Standalone JSON object
                r'\[.*\]'                  # JSON array
            ], 1):
                print(f"\n[DEBUG] Trying pattern {pattern_idx}: {pattern}")
                match = re.search(pattern, content, re.DOTALL)
                if match:
                    json_content = match.group(1) if pattern_idx == 1 else match.group(0)
                    print(f"[DEBUG] Found potential JSON content (original length: {len(json_content)})")

                    # Perform comment filtering
                    cleaned_content = remove_json_comments(json_content)
                    print(f"[DEBUG] Cleaned content length: {len(cleaned_content)}")

                    # Structural repair (including illegal escape fix)
                    repaired_content = FileProcessor.structural_repair(cleaned_content)
                    print(f"[DEBUG] Repaired content length: {len(repaired_content)}")

                    # Validate JSON
                    try:
                        print("[DEBUG] Validating JSON format...")
                        parsed = safe_json_parse(repaired_content)
                        print(f"[DEBUG] JSON validation successful! First 3 levels:\n{json.dumps(parsed, indent=2, ensure_ascii=False)[:500]}")
                        return parsed
                    except Exception as e:
                        print(f"[WARN] JSON validation failed: {str(e)}")
                        if isinstance(e, json.JSONDecodeError):
                            line_num = content.count('\n', 0, e.pos) + 1
                            col_num = e.pos - content.rfind('\n', 0, e.pos)
                            print(f"[DEBUG] Error position: Line {line_num} Column {col_num}")
                            print(f"[DEBUG] Error context:\n{content[max(0, e.pos-50):e.pos+50]}")
                        continue

            # Final fallback: try parsing entire file content (with comment filtering)
            print("\n[DEBUG] Trying to parse entire file content...")
            cleaned_content = remove_json_comments(content)
            repaired_content = FileProcessor.structural_repair(cleaned_content)
            return safe_json_parse(repaired_content)

        except Exception as e:
            print(f"[FATAL] {'Standard file' if is_std_file else 'Model output file'} loading failed: {str(e)}")
            print(f"[DEBUG] Error details: {type(e).__name__}: {str(e)}")
            if isinstance(e, json.JSONDecodeError):
                error_context = content[max(0, e.pos-50):e.pos+50]
                print(f"[DEBUG] Error context: {error_context}")
            raise

    @staticmethod
    def extract_pmid(file_path):
        """Extract PMID from filename"""
        return os.path.splitext(os.path.basename(file_path))[0].split('_')[0]

    @staticmethod
    def remove_json_comments(json_str):
        """Remove comments from JSON using regex"""
        pattern = r'(\s*//.*?\n)|(/\*.*?\*/)'
        return re.sub(pattern, '', json_str, flags=re.DOTALL)

    @staticmethod
    def safe_json_parse(json_str):
        """JSON parsing with automatic repair"""
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as e:
            fixed = re.sub(r',\s*}(?=\s*})', '}', json_str)
            fixed = re.sub(r'[\x00-\x1F\x7F]', '', fixed)
            try:
                return json.loads(fixed)
            except:
                import json5
                return json5.loads(fixed)


# Main execution flow
if __name__ == "__main__":
    # Configuration paths
    STD_DIR = r'.\ref_json'  # Standard JSON files directory
    MODEL_DIR = r'../06additional_data/result/01llama_70b_textRAG'
    OUTPUT_PATH = r'../06additional_data/01result_xlsx/local_01llama_70b_textRAGG.xlsx'

    # Initialize statistics variables
    total_files = 0
    success_pmids = []
    failed_pmids = {}

    print("="*50)
    print("Starting literature extraction results comparison...")
    print("="*50)

    # Get standard files list
    std_files = [f for f in os.listdir(STD_DIR) if f.endswith('.json')]
    total_files = len(std_files)
    print(f"Detected {total_files} documents to process")
    print()

    # Traverse standard files and perform comparison
    for std_file in std_files:
        pmid = os.path.splitext(std_file)[0]
        print(f"Processing PMID: {pmid}...")

        try:
            std_file_path = os.path.join(STD_DIR, std_file)
            model_json = None
            model_file_used = None

            # Try matching model file patterns (priority order)
            possible_patterns = [
                f'{pmid}.json',           # Pattern 1: Same name as standard file
                f'{pmid}_01.txt',         # Pattern 2: Original numbered format
                f'{pmid}_result.txt',     # Pattern 3: New format (result)
                f'{pmid}_*.txt'           # Pattern 4: Wildcard match for other formats
            ]

            # Try matching files by priority
            for pattern in possible_patterns:
                matched_files = [f for f in os.listdir(MODEL_DIR)
                                 if fnmatch.fnmatch(f, pattern)]

                if not matched_files:
                    continue

                # Prioritize exact matches
                target_file = None
                if pattern in [f'{pmid}.json', f'{pmid}_01.txt', f'{pmid}_result.txt']:
                    target_file = pattern if os.path.exists(os.path.join(MODEL_DIR, pattern)) else None
                else:
                    target_file = matched_files[0]

                if target_file:
                    model_path = os.path.join(MODEL_DIR, target_file)
                    try:
                        model_json = FileProcessor.load_json(model_path)
                        model_file_used = model_path
                        print(f"✅ Successfully loaded model output file: {target_file}")
                        break
                    except Exception as e:
                        print(f"⚠️ File {target_file} loading failed: {str(e)}")
                        continue

            # If no matching file found
            if not model_json:
                error_msg = "No matching model file found"
                print(f"⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Step 1: Data comparison
            print(f"\n🔍 Starting comparison for PMID {pmid}")
            try:
                std_json = FileProcessor.load_json(std_file_path, is_std_file=True)
                print(f"✅ Successfully loaded standard file")
            except Exception as e:
                error_msg = f"Standard file loading failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Perform comparison
            try:
                comparator = Step1_DataComparator(std_json)
                analysis_result = comparator.compare(model_json)
                print(f"✅ Data comparison completed")
            except Exception as e:
                error_msg = f"Data comparison failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg
                continue

            # Extract processing time
            with open(model_file_used, 'r', encoding='utf-8') as f:
                model_content = f.read()
            qa_time = TimeExtractor.extract(model_content)
            if qa_time:
                print(f"⏱️ Processing time: {qa_time} seconds")

            # Step 2: Create/load template
            if os.path.exists(OUTPUT_PATH):
                wb = load_workbook(OUTPUT_PATH)
            else:
                wb = Step2_TemplateBuilder.create_template(OUTPUT_PATH)

            # Step 3: Fill results
            try:
                Step3_ResultFiller.fill_results(wb, OUTPUT_PATH, pmid, analysis_result, qa_time)
                print(f"💾 Results saved to Excel")
                success_pmids.append(pmid)
            except Exception as e:
                error_msg = f"Result filling failed: {str(e)}"
                print(f"\n⛔ {error_msg}")
                failed_pmids[pmid] = error_msg

            print(f"✅ PMID {pmid} processing completed")
            print("="*50)

        except Exception as e:
            error_msg = f"Unknown error: {str(e)}"
            print(f"\n⛔ Unknown error occurred while processing PMID {pmid}: {str(e)}")
            failed_pmids[pmid] = error_msg

    # Output final statistics
    print("\n" + "="*50)
    print("📊 Processing Results Statistics")
    print("="*50)
    print(f"Total Documents: {total_files}")
    print(f"Successfully Processed: {len(success_pmids)} documents")
    print(f"Failed Processing: {len(failed_pmids)} documents")
    print(f"Success Rate: {len(success_pmids)/total_files*100:.1f}%")

    # Output failure details
    if failed_pmids:
        print("\n⛔ Failed PMIDs and Reasons:")
        for pmid, reason in failed_pmids.items():
            print(f"  {pmid}: {reason}")

    # Output success list
    if success_pmids:
        print("\n✅ Successfully Processed PMIDs:")
        print(", ".join(success_pmids))
    else:
        print("\n⛔ No PMIDs successfully processed")

    print("\nAll file processing completed!")